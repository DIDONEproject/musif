import copy
from multiprocessing import Lock
import os
from os import path

import musif.extract.features.ambitus as ambitus
import musif.extract.features.lyrics as lyrics
import openpyxl
from config import Configuration
from music21 import interval
from musif.common.constants import get_color
from musif.reports.constants import *
from musif.reports.utils import (adjust_excel_width_height,
                                 columns_alike_our_data)
from musif.reports.visualisations import bar_plot, box_plot, double_bar_plot, melody_bar_plot, pie_plot
from pandas.core.frame import DataFrame
from musif.reports.report_generation import excel_sheet

def Intervals(rows_groups: dict, not_used_cols: dict, factor, _cfg: Configuration, data: DataFrame, name: str, sorting_list: list, results_path: str, visualiser_lock: Lock, additional_info: list=[], groups: list=None):
    try:
        workbook = openpyxl.Workbook()
        all_columns = data.columns.tolist()
        general_cols = copy.deepcopy(not_used_cols)
        for row in rows_groups:
            if len(rows_groups[row][0]) == 0:
                general_cols.append(row)
            else:
                general_cols += rows_groups[row][0]

        # nombres de todos los intervalos
        third_columns_names_origin = set(all_columns) - set(general_cols)
        third_columns_names_origin = sort(
            third_columns_names_origin, sorting_list)
        third_columns_names = ['Total analysed'] + third_columns_names_origin

        # esta sheet va de sumar, así que en todas las columnas el cómputo que hay que hacer es sumar!
        computations = ["sum"]*len(third_columns_names)

        excel_sheet(workbook.create_sheet("Weighted"), third_columns_names, data, third_columns_names, computations, _cfg.sorting_lists,
                     groups=groups, average=True, last_column=True, last_column_average=False, additional_info=additional_info, ponderate=True)
        if factor>=1:
            excel_sheet(workbook.create_sheet("Horizontal Per"), third_columns_names, data, third_columns_names, computations, _cfg.sorting_lists,
                     groups=groups, per=True, average=True, last_column=True, last_column_average=False, additional_info=additional_info)


        if "Sheet" in workbook.get_sheet_names():  # Delete the default sheet
            std = workbook.get_sheet_by_name('Sheet')
            workbook.remove_sheet(std)
        adjust_excel_width_height(workbook)
            
        workbook.save(os.path.join(results_path, name))

        # with visualiser_lock:
        # VISUALISATIONS
        if 'Clefs' in name:
            title = 'Use of clefs in the vocal part'
        elif 'Intervals_absolute' in name:
            title = 'Presence of intervals (direction dismissed)'
        else:
            title = 'Presence of intervals (ascending and descending)'

        if groups:
            data_grouped = data.groupby(list(groups))
            for g, datag in data_grouped:
                result_visualisations = path.join(
                    results_path, 'visualisations', g)
                if not os.path.exists(result_visualisations):
                    os.mkdir(result_visualisations)
                name_bar = path.join(
                    result_visualisations, name.replace('.xlsx', IMAGE_EXTENSION))
                bar_plot(name_bar, datag, third_columns_names_origin,
                            'Intervals' if 'Clef' not in name else 'Clefs', title, second_title=str(g))
        elif factor == 1:
            for row in rows_groups:
                if row not in not_used_cols:
                    plot_name = name.replace(
                            '.xlsx', '') + '_Per_' + str(row.replace('Aria','').upper())
                    name_bar=path.join(results_path,'visualisations','Per_'+row.replace('Aria','').upper())
                    if not os.path.exists(name_bar):
                        os.makedirs(name_bar)

                    name_bar=path.join(name_bar,plot_name)

                    if len(rows_groups[row][0]) == 0:  # no sub-groups
                        data_grouped = data.groupby(row, sort=True)
                        if data_grouped:
                            bar_plot(name_bar + IMAGE_EXTENSION, data_grouped, third_columns_names_origin,
                                        'Intervals' + '\n' + str(row).replace('Aria','').upper() if 'Clef' not in name else 'Clefs' + str(row).replace('Aria','').upper(), title)
                    else: #subgroups
                        for i, subrow in enumerate(rows_groups[row][0]):
                            if subrow not in EXCEPTIONS:
                                data_grouped = data.groupby(subrow)
                                bar_plot(name_bar+'_'+subrow + IMAGE_EXTENSION, data_grouped, third_columns_names_origin,
                                        'Intervals' + str(row).replace('Aria','').upper() if 'Clef' not in name else 'Clefs' + str(row).replace('Aria','').upper(), title)

        else:
            name_bar = path.join(
                results_path, 'visualisations', name.replace('.xlsx', IMAGE_EXTENSION))
            bar_plot(name_bar, data, third_columns_names_origin,
                        'Intervals' if 'Clef' not in name else 'Clefs', title)
    except Exception as e:
        _cfg.write_logger.warn(get_color('WARNING')+'{}  Problem found: {}{}'.format(name, e, RESET_SEQ))


#########################################################
# Function to generate the reports file Intervals_types.xlsx  #
#########################################################

def Intervals_types(factor, _cfg: Configuration, data: DataFrame, results_path: str, name: str, visualiser_lock: Lock, groups=None, additional_info: list=[]):
    try:
        data.columns=[c.replace('Desc', 'Descending').replace('Asc', 'Ascending') for c in data.columns]
        workbook = openpyxl.Workbook()
        second_column_names = [("", 2), ("Leaps", 3), ("StepwiseMotion", 3)]
        second_column_names2 = [('', 1), ("Perfect", 3), ("Major", 3),
                                ("Minor", 3), ("Augmented", 3), ("Diminished", 3)]
        third_columns_names = ['Total analysed', "RepeatedNotes",
                               "Ascending", "Descending", "All", "Ascending", "Descending", "All"]
        third_columns_names2 = ['Total analysed', "Ascending", "Descending", "All", "Ascending", "Descending", "All",
                                "Ascending", "Descending", "All", "Ascending", "Descending", "All", "Ascending", "Descending", "All"]
        computations = ["sum"]*len(third_columns_names)
        computations2 = ['sum']*len(third_columns_names2)
        columns = columns_alike_our_data(
            third_columns_names, second_column_names)
        columns2 = columns_alike_our_data(
            third_columns_names2, second_column_names2)

        excel_sheet(workbook.create_sheet("Weighted"),
         columns, data, third_columns_names,
          computations, _cfg.sorting_lists, groups=groups, 
          last_column=True, last_column_average=False, second_columns=second_column_names,
           average=True,
                     columns2=columns2,  third_columns2=third_columns_names2, computations_columns2=computations2, second_columns2=second_column_names2, additional_info=additional_info, ponderate=True)
        if factor>=1:
            excel_sheet(workbook.create_sheet("Horizontal Per"), columns, data, third_columns_names, computations, _cfg.sorting_lists, groups=groups, second_columns=second_column_names, per=True, average=True, last_column=True, last_column_average=False,
                     columns2=columns2,  third_columns2=third_columns_names2, computations_columns2=computations2, second_columns2=second_column_names2, additional_info=additional_info)

        # borramos la sheet por defecto
        if "Sheet" in workbook.get_sheet_names():
            std = workbook.get_sheet_by_name('Sheet')
            workbook.remove_sheet(std)
        adjust_excel_width_height(workbook)
        workbook.save(os.path.join(results_path, name))

        # with visualiser_lock:
        # VISUALISATIONS
        if groups:
            data_grouped = data.groupby(list(groups))
            for g, datag in data_grouped:
                result_visualisations = path.join(
                    results_path, 'visualisations', g)
                if not os.path.exists(result_visualisations):
                    os.mkdir(result_visualisations)

                name_cakes = path.join(result_visualisations,
                                    name.replace('.xlsx',  '') + '_AD.png')
                pie_plot(name_cakes, datag, second_title=str(g))
                name_bars = path.join(result_visualisations,
                                    name.replace('.xlsx',  IMAGE_EXTENSION))
                double_bar_plot(name_bars, data, second_title=str(g))

        elif factor == 1:
            # groups = [i for i in rows_groups]
            for row in rows_groups:
                name_folder =path.join(results_path,'visualisations','Per_'+row.replace('Aria','').upper())

                name_cakes = name.replace(
                            '.xlsx', '') + '_Per_' + str(row.replace('Aria','').upper())  + '_AD.png'
                name_bars = path.join(results_path, 'visualisations',
                                name.replace('.xlsx',  '').replace('1_factor','') + '_Per_' + str(row.replace('Aria','').upper()) + IMAGE_EXTENSION)
              
                if not os.path.exists(name_folder):
                    os.makedirs(name_folder)

                name_cakes=path.join(name_folder,name_cakes)
                name_bars=path.join(name_folder,name_bars)

                
                if row not in not_used_cols:
                    if len(rows_groups[row][0]) == 0:  # no sub-groups
                        data_grouped = data.groupby(row, sort=True)
                        if data_grouped:
                            # melody_bar_plot(name_bar, data_grouped, columns_visualisation, second_title='Per ' + str(subrow.replace('Aria','').upper()))

                            pie_plot(name_cakes, data_grouped, second_title='Per ' + str(row.replace('Aria','').upper()))
                            double_bar_plot(name_bars, data_grouped, 'Per ' + str(row.replace('Aria','').upper()))

                    else: #subgroups
                            for i, subrow in enumerate(rows_groups[row][0]):
                                if subrow not in EXCEPTIONS:
                                    name_cakes = path.join(results_path, 'visualisations',
                                    name.replace('.xlsx', '') + '_Per_' + str(row.upper()) + '_' + str(subrow) + '_AD.png')
                                    name_bars = path.join(results_path, 'visualisations',
                                    name.replace('.xlsx',  '') + '_Per_' + str(row.upper()) + '_' + str(subrow) + IMAGE_EXTENSION)

                                    data_grouped = data.groupby(subrow)
                                    pie_plot(name_cakes, data_grouped, second_title='Per ' + str(row.replace('Aria','').upper()))
                                    double_bar_plot(name_bars, data_grouped, 'Per ' + str(row.replace('Aria','').upper()))

        else:
            name_cakes = path.join(results_path, 'visualisations',
                                name.replace('.xlsx', '') + '_AD.png')
            pie_plot(name_cakes, data)
            name_bars = path.join(results_path, 'visualisations',
                                name.replace('.xlsx',  IMAGE_EXTENSION))
            double_bar_plot(name_bars, data)
    except Exception as e:
        _cfg.write_logger.warn(get_color('WARNING')+'{}  Problem found: {}{}'.format(name, e, RESET_SEQ))
