import copy
from musif.reports.visualisations import bar_plot_extended, line_plot_extended
import os
from multiprocessing import Lock
from os import path
import numpy as np
from musif.common.constants import get_color

import pandas as pd
from musif.common.sort import sort

import musif.extract.features.ambitus as ambitus
import musif.extract.features.lyrics as lyrics
import openpyxl
from config import Configuration
from music21 import interval
from musif.reports.constants import *
from musif.reports.utils import adjust_excel_width_height, excel_sheet
from pandas.core.frame import DataFrame

### COMMON INSTRUMENTS tasks ###

### Function that prints densities report task ###

def Densities(rows_groups: dict, not_used_cols: dict, factor, _cfg: Configuration, data: DataFrame, results_path: str, name: str, visualiser_lock: Lock, groups: list=None, additional_info: list=[]):
    try:
        workbook = openpyxl.Workbook()

        # Splitting the dataframes to reorder them
        data_general = data[metadata_columns + ['Total analysed']].copy()
        data = data[set(data.columns).difference(metadata_columns)]
        data_general = data_general.dropna(how='all', axis=1)
        data = data.dropna(how='all', axis=1)
        density_df =data[[i for i in data.columns if i.endswith('SoundingDensity')]]
        notes_and_measures =data[[i for i in data.columns if i.endswith(
            'Measures') or i.endswith('Notes') or i.endswith('Mean') or i == 'NumberOfBeats']]

        density_df.columns = [i.replace('_','').replace('Part', '').replace(
            'Sounding', '').replace('Density', '').replace('Family', '').replace('Sound', '').replace('Notes', '') for i in density_df.columns]
        notes_and_measures.columns = [i.replace('_','').replace('Part', '').replace('SoundingMeasures', 'Measures').replace(
            'Sound', '').replace('Notes', '').replace('Mean', '').replace('Family', '') for i in notes_and_measures.columns]

        cols = sort(density_df.columns.tolist(), [
                    i.capitalize() for i in _cfg.sorting_lists['InstrumentSorting']])
        
        density_df = density_df[cols]
        third_columns_names = density_df.columns.to_list()

        second_column_names = [("", 1), ("Density", len(third_columns_names))]
        third_columns_names.insert(0, 'Total analysed')
        data = pd.concat([data_general, density_df], axis=1)
        data_total = pd.concat([data_general, notes_and_measures], axis=1)

        computations = ["sum"] + ["mean"] * (len(third_columns_names)-1)
        computations2 = ["sum"] + ["mean_density"] * \
            (len(third_columns_names)-1)
        columns = third_columns_names
        excel_sheet(workbook.create_sheet("Weighted"), columns, data, third_columns_names, computations, _cfg.sorting_lists, groups=groups, last_column=True,
                     last_column_average=True, second_columns=second_column_names, average=True, additional_info=additional_info, ponderate=False)
        if factor >=1:
            excel_sheet(workbook.create_sheet("Horizontal"), columns, data_total, third_columns_names, computations2,  _cfg.sorting_lists, groups=groups,
                     second_columns=second_column_names, per=False, average=True, last_column=True, last_column_average=True, additional_info=additional_info)

        # Deleting default sheet
        if "Sheet" in workbook.get_sheet_names():
            std = workbook.get_sheet_by_name('Sheet')
            workbook.remove_sheet(std)
        adjust_excel_width_height(workbook)
        workbook.save(os.path.join(results_path, name))

        # with visualiser_lock: #Apply when threads are usedwith visualizer_lock=threading.Lock()
        columns.remove('Total analysed')
        title = 'Instrumental densities'
        # VISUALISATIONS
        if groups:
            data_grouped = data.groupby(list(groups))
            for g, datag in data_grouped:
                result_visualisations = results_path + \
                    '\\visualisations\\' + str(g.replace('/', '_'))
                if not os.path.exists(result_visualisations):
                    os.mkdir(result_visualisations)
                name_bar = result_visualisations + \
                    '\\' + name.replace('.xlsx', IMAGE_EXTENSION)
                bar_plot_extended(name_bar, datag, columns, 'Density',
                                  'Density', title, second_title=str(g))

        elif factor == 1:  # 
            groups = [i for i in rows_groups]
            for row in rows_groups:
                plot_name = name.replace(
                    '.xlsx', '') + '_Per_' + str(row.upper()) + IMAGE_EXTENSION
                name_bar = results_path + '\\visualisations\\' + plot_name
                if row not in not_used_cols:
                    if len(rows_groups[row][0]) == 0:  # no sub-groups
                        data_grouped = data.groupby(row, sort=True)
                        if data_grouped:
                            line_plot_extended(
                            name_bar, data_grouped, columns, 'Instrument', 'Density', title, second_title='Per ' + str(row))
                    else:
                        for i, subrow in enumerate(rows_groups[row][0]):
                            if subrow not in EXCEPTIONS:
                                plot_name = name.replace(
                                    '.xlsx', '') + '_Per_' + str(row.upper()) + '_' + str(subrow) + IMAGE_EXTENSION
                                name_bar = results_path + '\\visualisations\\' + plot_name
                                data_grouped = data.groupby(subrow)
                                line_plot_extended(
                                    name_bar, data_grouped, columns, 'Instrument', 'Density', title, second_title= 'Per ' + str(subrow))
        else:
            for instr in third_columns_names:
                columns=data['AriaName']
                name_bar = results_path + '\\visualisations\\' + instr + \
                    name.replace('.xlsx', IMAGE_EXTENSION)
                bar_plot_extended(name_bar, data, columns,
                                'Density', 'Density', title + ' ' + instr, instr=instr)
    except Exception as e:
        _cfg.write_logger.warn(get_color('WARNING')+'{}  Problem found: {}{}'.format(name, e, RESET_SEQ))

### Function that prints textures report task ###

def Textures(rows_groups: dict, not_used_cols: dict, factor, _cfg: Configuration, data: DataFrame, results_path: str, name: str, visualiser_lock: Lock, groups: list=None, additional_info: list=[]):
    try:
        workbook = openpyxl.Workbook()
        # Splitting the dataframes to reorder them
        data_general = data[metadata_columns + ['Total analysed']].copy()
        data_general = data_general.dropna(how='all', axis=1)
        notes_df = data[[i for i in data.columns if i.endswith('Notes') or i.endswith('Mean')]]
        textures_df = data[[c for c in data.columns if c.endswith('Texture')]]
        third_columns_names = []
        # Pre-treating data columns
        notes_df.columns = [i.replace('_', '').replace('Sound','').replace('Family','').replace('Part','').replace('Mean', '') for i in notes_df.columns]
        textures_df.columns = [i.replace('__Texture','').replace('__', '/').replace('Part', '') for i in textures_df.columns]
        
        for column in textures_df.columns:
            if column not in _cfg.sorting_lists["TexturesSorting"]:
                textures_df.drop([column], axis=1, inplace=True)
            else:
                notes_df[column]=textures_df[column]

        third_columns_names = _cfg.sorting_lists['TexturesSorting']

        for col in third_columns_names:
            if col not in textures_df.columns:
                textures_df[col]=np.nan
                notes_df[col]=np.nan
                notes_df[col.split('/')[0]+'_Notes']=np.nan

        cols = sort(textures_df.columns.tolist(), _cfg.sorting_lists['TexturesSorting'])  

        textures_df = textures_df[cols]
        
        # second_column_names = [("", 1), ("Textures", len(third_columns_names))]
        third_columns_names.insert(0, 'Total analysed')
        second_column_names = [
            ("", 1), ("Texture", len(third_columns_names))]

        computations = ["sum"] + ["mean"] * (len(third_columns_names)-1)
        computations2 = ["sum"] + ["mean_texture"] * \
            (len(third_columns_names)-1)
        columns = third_columns_names

        data = pd.concat([data_general, textures_df], axis=1)
        notes_df = pd.concat([data_general, notes_df], axis=1)

        excel_sheet(workbook.create_sheet("Weighted_textures"), columns, data, third_columns_names, computations, _cfg.sorting_lists, groups=groups,
                     last_column=True, last_column_average=True, second_columns=second_column_names, average=True, additional_info=additional_info, ponderate=False)
        if factor >=1:
            excel_sheet(workbook.create_sheet("Horizontal_textures"), columns, notes_df, third_columns_names, computations2,   _cfg.sorting_lists, groups=groups,
                     second_columns=second_column_names, per=False, average=True, last_column=True, last_column_average=True, additional_info=additional_info)

        # borramos la sheet por defecto
        if "Sheet" in workbook.get_sheet_names():
            std = workbook.get_sheet_by_name('Sheet')
            workbook.remove_sheet(std)
            
        adjust_excel_width_height(workbook)
        workbook.save(os.path.join(results_path, name))

        # with visualiser_lock:
        title = 'Textures'
        columns.remove('Total analysed')
        # VISUALISATIONS
        if groups:
            data_grouped = data.groupby(list(groups))
            for g, datag in data_grouped:
                result_visualisations = results_path + \
                    '\\visualisations\\' + str(g)
                if not os.path.exists(result_visualisations):
                    os.mkdir(result_visualisations)
                name_bar = result_visualisations + \
                    '\\' + name.replace('.xlsx', IMAGE_EXTENSION)
                bar_plot_extended(
                    name_bar, datag, columns, 'Instrumental Textures', title, second_title=str(g))

        elif factor == 1:
            groups = [i for i in rows_groups]
            for row in rows_groups:
                plot_name = name.replace(
                    '.xlsx', '') + '_Per_' + str(row.upper()) + IMAGE_EXTENSION
                name_bar = results_path + '\\visualisations\\' + plot_name
                if row not in not_used_cols:
                    if len(rows_groups[row][0]) == 0:  # no sub-groups
                        data_grouped = data.groupby(row, sort=True)
                        line_plot_extended(
                            name_bar, data_grouped, columns, 'Texture', 'Ratio', title, second_title='Per ' + str(row))
                    else:
                        for i, subrow in enumerate(rows_groups[row][0]):
                            if subrow not in EXCEPTIONS:
                                plot_name = name.replace(
                                    '.xlsx', '') + '_Per_' + str(row.upper()) + '_' + str(subrow) + IMAGE_EXTENSION
                                name_bar = results_path + '\\visualisations\\' + plot_name
                                data_grouped = data.groupby(subrow)
                                line_plot_extended(
                                    name_bar, data_grouped, columns, 'Texture', 'Ratio', title, second_title='Per ' + str(subrow))
        else:
            for instr in third_columns_names:
                if not all(data[instr].isnull()):
                    columns=data['AriaName']
                    name_bar = results_path + '\\visualisations\\' + instr.replace('/', '_') + \
                        name.replace('.xlsx', IMAGE_EXTENSION)
                    bar_plot_extended(name_bar, data, columns,
                                        'Instrumental Textures', 'Ratio', title + ' ' +instr, instr=instr)
    except Exception as e:
            _cfg.write_logger.info('{}  Problem found: {}'.format(name, e))
