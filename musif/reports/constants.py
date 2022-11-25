import openpyxl
from openpyxl.styles.fonts import Font
from musif.extract.features import ambitus
from musif.extract.features import interval

from musif.extract.basic_modules.scoring.constants import (
    FAMILY_SCORING,
    SCORING,
    VOICES,
)
from musif.extract.features.tempo.constants import TEMPO_GROUPED_1, TEMPO_GROUPED_2
from musif.extract.basic_modules.metadata.constants import *

ACT = "Act"
# ACTANDSCENE = 'ActAndScene'
ARIA_ID = "AriaId"
ARIA_LABEL = "AriaLabel"
# CHARACTER = 'Character'
# CITY = 'City'
CLEF1 = "Clef1"
CLEF2 = "Clef2"
CLEF3 = "Clef3"
COMPOSER = "Composer"
DATE = "Date"
DECADE = "Decade"
DRAMA = "Drama"
FINAL = "Final"
# FORM = 'Form'
# GENDER ='Gender'
# GEOGRAPHY='Geography'
KEY = "Key"
KEY_SIGNATURE_TYPE = "KeySignatureType"
KEYSIGNATURE = "KeySignature"
LIBRETTIST = "Librettist"
METRE = "Metre"
MODE = "Mode"
NAME = "FileName"
OPERA = "AriaOpera"
ROLE = "RoleType"
SCENE = "Scene"
TEMPO = "Tempo"
# TERRITORY = 'Territory'
TIMESIGNATURE = "TimeSignature"
TIMESIGNATUREGROUPED = "TimeSignatureGrouped"
TITLE = "AriaName"
TOTAL_ANALYSED = "Total analysed"
YEAR = "Year"

metadata_columns = [
    OPERA,
    ARIA_LABEL,
    ARIA_ID,
    TITLE,
    COMPOSER,
    YEAR,
    DECADE,
    ACT,
    SCENE,
    ACTANDSCENE,
    NAME,
    LIBRETTIST,
    FORM,
    CHARACTER,
    GENDER,
    ROLE,
    ARIA_CITY,
    TERRITORY,
    CLEF1,
    CLEF2,
    CLEF3,
    KEY,
    KEYSIGNATURE,
    KEY_SIGNATURE_TYPE,
    MODE,
    TEMPO,
    TIMESIGNATURE,
    TIMESIGNATUREGROUPED,
    TEMPO_GROUPED_1,
    TEMPO_GROUPED_2,
    SCORING,
    FAMILY_SCORING,
]


def get_melody_list(catch):
    return [
        catch + interval.constants.INTERVALLIC_MEAN,
        catch + interval.constants.INTERVALLIC_STD,
        catch + interval.constants.ABSOLUTE_INTERVALLIC_MEAN,
        catch + interval.constants.ABSOLUTE_INTERVALLIC_STD,
        catch + interval.constants.TRIMMED_ABSOLUTE_INTERVALLIC_MEAN,
        catch + interval.constants.TRIMMED_ABSOLUTE_INTERVALLIC_STD,
        catch + interval.constants.TRIMMED_INTERVALLIC_STD,
        catch + interval.constants.TRIMMED_INTERVALLIC_MEAN,
        catch + interval.constants.ABSOLUTE_INTERVALLIC_TRIM_DIFF,
        catch + interval.constants.ABSOLUTE_INTERVALLIC_TRIM_RATIO,
        catch + interval.constants.LARGEST_ABSOLUTE_SEMITONES_ASC,
        catch + interval.constants.LARGEST_INTERVAL_ASC,
        catch + interval.constants.LARGEST_SEMITONES_DESC,
        catch + interval.constants.LARGEST_SEMITONES_ASC,
        catch + interval.constants.LARGEST_INTERVAL_DESC,
        catch + ambitus.constants.LOWEST_NOTE,
        catch + ambitus.constants.LOWEST_NOTE_INDEX,
        catch + ambitus.constants.HIGHEST_NOTE,
        catch + ambitus.constants.HIGHEST_NOTE_INDEX,
        catch + interval.constants.LARGEST_INTERVAL_ALL,
        catch + interval.constants.LARGEST_SEMITONES_ALL,
        catch + interval.constants.SMALLEST_INTERVAL_ALL,
        catch + interval.constants.SMALLEST_SEMITONES_ALL,
        catch + interval.constants.MEAN_INTERVAL,
        catch + interval.constants.DESCENDING_INTERVALLIC_MEAN,
        catch + interval.constants.ASCENDING_INTERVALLIC_MEAN,
    ]


not_used_cols = [ARIA_ID, SCORING, TOTAL_ANALYSED, CLEF2, CLEF3]

EXCEPTIONS = [ROLE, KEYSIGNATURE, TEMPO, YEAR, ARIA_CITY, SCENE, NAME]

alfa = "abcdefghijklmnopqrstuvwxyz"
COMMON_DF = "common_df"
VISUALIZATIONS = "visualizations"
forbiden_groups = {
    OPERA: [OPERA],
    ARIA_LABEL: [OPERA, ARIA_LABEL],
    TITLE: [TITLE, OPERA],
    COMPOSER: [COMPOSER],
    NAME: [NAME],
    YEAR: [YEAR, DECADE],
    DECADE: [DECADE],
    ARIA_CITY: [ARIA_CITY, TERRITORY],
    TERRITORY: [TERRITORY],
    ACT: [ACT, ACTANDSCENE],
    SCENE: [SCENE, ACTANDSCENE],
    ACTANDSCENE: [ACT, SCENE, ACTANDSCENE],
    ROLE: [ROLE, GENDER],
    GENDER: [GENDER],
    FORM: [FORM],
    CLEF1: [CLEF1],
    CLEF2: [CLEF2],
    CLEF3: [CLEF3],
    KEY: [FORM, MODE, FINAL, KEYSIGNATURE, KEY_SIGNATURE_TYPE],
    MODE: [MODE, FINAL],
    FINAL: [FINAL, KEY, KEYSIGNATURE],
    KEYSIGNATURE: [KEY, FINAL, KEYSIGNATURE, KEY_SIGNATURE_TYPE],
    KEY_SIGNATURE_TYPE: [KEY_SIGNATURE_TYPE],
    TIMESIGNATURE: [TIMESIGNATURE, TIMESIGNATUREGROUPED],
    TIMESIGNATUREGROUPED: [TIMESIGNATUREGROUPED],
    TEMPO: [TEMPO, TEMPO_GROUPED_1, TEMPO_GROUPED_2],
    TEMPO_GROUPED_1: [TEMPO_GROUPED_1, TEMPO_GROUPED_2],
    TEMPO_GROUPED_2: [TEMPO_GROUPED_2],
    "AbrScoring": ["AbrScoring", "RealScoringGrouped"],
    "RealScoringGrouped": ["RealScoringGrouped"],
}

rows_groups = {
    OPERA: ([], "Alphabetic"),
    ARIA_LABEL: ([], "Alphabetic"),
    TITLE: ([], "Alphabetic"),
    COMPOSER: ([], "Alphabetic"),
    NAME: ([], "Alphabetic"),
    DATE: (
        [
            YEAR,
            DECADE,
        ],
        ["Alphabetic", "Alphabetic"],
    ),
    GEOGRAPHY: ([ARIA_CITY, TERRITORY], ["Alphabetic", "Alphabetic"]),
    DRAMA: ([ACT, SCENE, ACTANDSCENE], ["Alphabetic", "Alphabetic", "Alphabetic"]),
    CHARACTER: (
        [CHARACTER, ROLE, GENDER],
        ["CharacterSorting", "Alphabetic", "Alphabetic"],
    ),
    FORM: ([], "FormSorting"),
    CLEF1: ([], "Alphabetic"),
    LIBRETTIST: ([], "Alphabetic"),
    KEY: (
        [KEY, MODE, KEYSIGNATURE, KEY_SIGNATURE_TYPE],
        [
            "KeySorting",
            "Alphabetic",
            "KeySignatureSorting",
            "KeySignatureGroupedSorted",
        ],
    ),
    METRE: (
        [TIMESIGNATURE, TIMESIGNATUREGROUPED],
        ["TimeSignatureSorting", "Alphabetic"],
    ),
    TEMPO: (
        [TEMPO, TEMPO_GROUPED_1, TEMPO_GROUPED_2],
        ["TempoSorting", "TempoGroupedSorting1", "TempoGroupedSorting2"],
    ),
    SCORING: (
        [
            SCORING,
            FAMILY_SCORING
            # Antes Scoring Sorting
        ],
        ["InstrumentSorting", "ScoringFamilySorting"],
    ),
}


YELLOWFILL = openpyxl.styles.PatternFill(
    start_color="F9E220", end_color="F9E220", fill_type="solid"
)
greenFill = openpyxl.styles.PatternFill(
    start_color="98E891", end_color="98E891", fill_type="solid"
)
orangeFill = openpyxl.styles.PatternFill(
    start_color="EE6513", end_color="EE6513", fill_type="solid"
)

titles_second_Fill = openpyxl.styles.PatternFill(
    start_color="FA9455", end_color="FA9455", fill_type="solid"
)

titles1Fill = openpyxl.styles.PatternFill(
    start_color="F97626", end_color="F97626", fill_type="solid"
)

titles2Fill = openpyxl.styles.PatternFill(
    start_color="93d3fb", end_color="93d3fb", fill_type="solid"
)

titles3Fill = openpyxl.styles.PatternFill(
    start_color="FBBA93", end_color="FBBA93", fill_type="solid"
)

titles4Fill = openpyxl.styles.PatternFill(
    start_color="00FFFF", end_color="00FFFF", fill_type="solid"
)

titles5Fill = openpyxl.styles.PatternFill(
    start_color="C0C0C0", end_color="C0C0C0", fill_type="solid"
)


fills_list = [
    titles1Fill,
    titles3Fill,
    titles4Fill,
    titles5Fill,
    titles1Fill,
    titles3Fill,
    titles4Fill,
    titles5Fill,
    titles1Fill,
    titles3Fill,
    titles4Fill,
    titles5Fill,
    titles1Fill,
    titles3Fill,
    titles4Fill,
    titles5Fill,
]

factors_Fill = [
    openpyxl.styles.PatternFill(
        start_color="06CAFF", end_color="06CAFF", fill_type="solid"
    ),
    openpyxl.styles.PatternFill(
        start_color="18ADD5", end_color="18ADD5", fill_type="solid"
    ),
    openpyxl.styles.PatternFill(
        start_color="1B94B4", end_color="1B94B4", fill_type="solid"
    ),
    openpyxl.styles.PatternFill(
        start_color="11718A", end_color="11718A", fill_type="solid"
    ),
]

center = openpyxl.styles.Alignment(horizontal="center")
BOLD = Font(size=12, bold=True)
FONT = Font(size=12)
FONT_TITLE = Font(size=12, bold=True, name="Garamond")

WIDE = 1.5
NORMAL_WIDTH = 1
NARROW = 0.5

IMAGE_EXTENSION = ".png"
