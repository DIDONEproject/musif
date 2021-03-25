import openpyxl
from .constants import not_used_cols
from .constants import rows_groups

########################################################################
# Function that finds the propper name used in our intermediate files  #
########################################################################


def columns_alike_our_data(third_columns_names, second_column_names, first_column_names=None):
    columns = []
    counter_first = 0
    sub_counter_first = 0
    counter_second = 0
    sub_counter_second = 0
    for c in third_columns_names:
        if first_column_names:
            cn = first_column_names[counter_first][0] + \
                second_column_names[counter_second][0] + c
            sub_counter_first += 1
            if sub_counter_first >= first_column_names[counter_first][1]:
                sub_counter_first = 0
                counter_first += 1
        else:
            cn = second_column_names[counter_second][0] + c
        sub_counter_second += 1
        if sub_counter_second >= second_column_names[counter_second][1]:
            sub_counter_second = 0
            counter_second += 1
        columns.append(cn)
    return columns

##########################################################################################################
# Function in charge of printting the data, the arguments are the same as the explained in hoja_iValues  #
##########################################################################################################


def row_iteration(hoja, columns, row_number, column_number, data, third_columns, computations_columns, sorting_lists, group=None, first_columns=None, second_columns=None, per=False, average=False, last_column=False, last_column_average=False,
                  columns2=None, data2=None, third_columns2=None, computations_columns2=None, first_columns2=None, second_columns2=None, additional_info=[], ponderate=False):
    all_columns = list(data.columns)
    for row in rows_groups:  # Geography, Dramma, Opera, Aria, Label, Composer...
        if row in all_columns or any(sub in all_columns for sub in rows_groups[row][0]):
            forbiden = []
            if group != None:
                forbiden = [forbiden_groups[group[i]]
                            for i in range(len(group))]
                forbiden = [item for sublist in forbiden for item in sublist]
            if group == None or row not in forbiden:
                # 1. Write the Title in Yellow
                hoja.cell(row_number, column_number).value = "Per " + row
                hoja.cell(row_number, column_number).fill = yellowFill
                row_number += 1
                sorting = rows_groups[row][1]
                # 2. Write the information depending on the subgroups (ex: Geography -> City, Country)
                if len(rows_groups[row][0]) == 0:  # No subgroups
                    starting_row = row_number
                    # Sort the dataframe based on the json sorting_lists in Json_extra
                    data = sort_dataframe(data, row, sorting_lists, sorting)
                    groups_add_info, add_info = get_groups_add_info(
                        data, row, additional_info)
                    row_number, last_column_used = print_groups(hoja, data.groupby(groups_add_info, sort=False), row_number, column_number, columns, third_columns, computations_columns, first_columns, second_columns, per=per,
                                                                average=average, last_column=last_column, last_column_average=last_column_average, additional_info=add_info, ponderate=ponderate, not_grouped_df=(groups_add_info, data[groups_add_info + columns]))
                    if columns2 != None:  # Second subgroup
                        groups_add_info, add_info = get_groups_add_info(
                            data, row, additional_info)
                        if data2 is not None:
                            data2 = sort_dataframe(
                                data2, row, sorting_lists, sorting)
                        _, _ = print_groups(hoja, data.groupby(groups_add_info, sort=False) if data2 is None else data2.groupby(groups_add_info, sort=False), starting_row, last_column_used + 1, columns2, third_columns2, computations_columns2, first_columns2,
                                            second_columns2, per=per, average=average, last_column=last_column, last_column_average=last_column_average, additional_info=add_info, ponderate=ponderate, not_grouped_df=(groups_add_info, data[groups_add_info + columns]))
                else:  # has subgroups, ex: row = Date, subgroups: Year
                    for i, subrows in enumerate(rows_groups[row][0]):
                        if (subrows == None or subrows not in forbiden) and subrows in all_columns:
                            if "Tempo" in subrows:
                                data[subrows] = data[subrows].fillna('')
                            starting_row = row_number
                            sort_method = sorting[i]
                            hoja.cell(
                                row_number, column_number).value = subrows
                            hoja.cell(
                                row_number, column_number).fill = greenFill
                            data = sort_dataframe(
                                data, subrows, sorting_lists, sort_method)

                            groups_add_info, add_info = get_groups_add_info(
                                data, subrows, additional_info)
                            row_number, last_column_used = print_groups(hoja, data.groupby(groups_add_info, sort=False), row_number, column_number, columns, third_columns, computations_columns, first_columns, second_columns, per=per,
                                                                        average=average, last_column=last_column, last_column_average=last_column_average, additional_info=add_info, ponderate=ponderate, not_grouped_df=(groups_add_info, data[groups_add_info + columns]))
                            if columns2 != None:  # Second subgroup
                                if "Tempo" in subrows and data2 is not None:
                                    data2[subrows] = data2[subrows].fillna('')
                                if data2 is not None:
                                    data2 = sort_dataframe(
                                        data2, subrows, sorting_lists, sort_method)
                                groups_add_info, add_info = get_groups_add_info(
                                    data, subrows, additional_info)
                                _, _ = print_groups(hoja, data.groupby(groups_add_info, sort=False) if data2 is None else data2.groupby(groups_add_info, sort=False), starting_row, last_column_used + 1, columns2, third_columns2, computations_columns2, first_columns2,
                                                    second_columns2, per=per, average=average, last_column=last_column, last_column_average=last_column_average, additional_info=add_info, ponderate=ponderate, not_grouped_df=(groups_add_info, data[groups_add_info + columns]))

                            row_number += 1
                row_number += 1
    return row_number


def hoja_iValues(hoja, columns, data, third_columns, computations_columns, sorting_lists, groups=None, first_columns=None, second_columns=None, per=False, average=False, last_column=False, last_column_average=False,
                 columns2=None, data2=None, third_columns2=None, computations_columns2=None, first_columns2=None, second_columns2=None, additional_info=[], ponderate=False):

    row_number = 1  # we start writing in row 1
    column_number = 1
    if groups == None:
        row_iteration(hoja, columns, row_number, column_number, data, third_columns, computations_columns, sorting_lists, first_columns=first_columns, second_columns=second_columns, per=per,
                      average=average, last_column=last_column, last_column_average=last_column_average, columns2=columns2, data2=data2, third_columns2=third_columns2, computations_columns2=computations_columns2, first_columns2=first_columns2, second_columns2=second_columns2, additional_info=additional_info, ponderate=ponderate)
    else:
        # we may be grouping by more than 2 factors
        data_grouped = data.groupby(list(groups))

        last_printed = {i: ('', 0) for i in range(len(groups))}
        for group, group_data in data_grouped:
            cnumber = column_number
            group = [group] if type(group) != tuple else group
            for i, g in enumerate(group):
                if last_printed[i][0] != g:
                    hoja.cell(row_number, cnumber).value = g
                    hoja.cell(row_number, cnumber).fill = factors_Fill[i]
                    hoja.cell(row_number, cnumber).font = bold
                    counter_g = data[groups[i]].tolist().count(g)
                    hoja.cell(row_number, cnumber + 1).value = counter_g
                    if last_printed[i][0] != '':
                        hoja.merge_cells(
                            start_row=last_printed[i][1], start_column=i + 1, end_row=row_number - 2, end_column=i + 1)
                        hoja.cell(last_printed[i][1],
                                  i + 1).fill = factors_Fill[i]

                    last_printed[i] = (g, row_number + 1)

                row_number += 1
                cnumber += 1
            data2_grouped = None
            if data2 is not None:
                data2_grouped = data2.groupby(list(groups)).get_group(
                    group if len(group) > 1 else group[0])
            rn = row_iteration(hoja, columns, row_number, cnumber, group_data, third_columns, computations_columns, sorting_lists, group=groups, first_columns=first_columns, second_columns=second_columns, per=per,
                               average=average, last_column=last_column, last_column_average=last_column_average, columns2=columns2, data2=data2_grouped, third_columns2=third_columns2, computations_columns2=computations_columns2, first_columns2=first_columns2, second_columns2=second_columns2, additional_info=additional_info, ponderate=ponderate)
            row_number = rn
        # merge last cells
        for i, g in enumerate(group):
            if last_printed[i][0] == g:
                hoja.merge_cells(
                    start_row=last_printed[i][1], start_column=i + 1, end_row=row_number - 2, end_column=i + 1)
                hoja.cell(last_printed[i][1],  i + 1).fill = factors_Fill[i]

########################################################################
# Function ment to write the iValues excel
# -------
# data: all_info dataframe
# results_path: where to store the data
# name: name that the excel will take
# sorting lists: lists that will be used for sorting the results
# visualiser_lock: lock used to avoid deadlocks, as matplotlib is not thread safe
# additional info: columns additional to each
# remove_columns: used for factor 0, to avoid showing unusefull information
# groups: used for factor > 1
########################################################################


def iValues(data, results_path, name, sorting_lists, visualiser_lock, additional_info=[], remove_columns=False, groups=None):
    try:
        workbook = openpyxl.Workbook()

        # HOJA 1: STATISTICAL_VALUES
        column_names = ["Total analysed", "Intervallic ratio", "Trimmed intervallic ratio", "dif. Trimmed",
                        "% Trimmed", "Absolute intervallic ratio", "Std", "Absolute Std", 'Syllabic ratio']
        # HAREMOS LA MEDIA DE TODAS LAS COLUMNAS
        computations = ['sum'] + ["mean"]*(len(column_names) - 1)
        hoja_iValues(workbook.create_sheet("Statistical_values"), column_names, data, column_names, computations,
                     sorting_lists, groups=groups, average=True, additional_info=additional_info, ponderate=True)

        # HOJA 2: AMBITUS
        first_column_names = [("", 1), ("Lowest", 2), ("Highest", 2), ("Lowest", 2), ("Highest", 2), (
            "Ambitus", 8)] if not remove_columns else [("", 1), ("Lowest", 2), ("Highest", 2), ("Ambitus", 2)]

        second_column_names = [("", 5), ("Mean", 2), ("Mean", 2), ("Largest", 2), ("Smallest", 2), (
            "Absolute", 2), ("Mean", 2)] if not remove_columns else [("", 5), ("Largest", 2)]

        third_columns_names = ["Total analysed", "Note", "Index", "Note", "Index", "Note", "Index", "Note", "Index", "Semitones", "Interval", "Semitones", "Interval",
                               "Semitones", "Interval", "Semitones", "Interval"] if not remove_columns else ["Total analysed", "Note", "Index", "Note", "Index", "Semitones", "Interval"]

        computations = ["sum", "minNote", "min", "maxNote", "max", 'meanNote', 'mean', 'meanNote', 'mean', 'max', "maxInterval", 'min', "minInterval", 'absolute',
                        'absoluteInterval', "meanSemitones", "meanInterval"] if not remove_columns else ["sum", "minNote", "min", "maxNote", "max", 'max', "maxInterval"]

        columns = columns_alike_our_data(
            third_columns_names, second_column_names, first_column_names)

        hoja_iValues(workbook.create_sheet("Ambitus"), columns, data, third_columns_names, computations, sorting_lists, groups=groups,
                     first_columns=first_column_names, second_columns=second_column_names, average=True, additional_info=additional_info)

        # HOJA 3: LARGEST_LEAPS
        second_column_names = [("", 1), ("Ascending", 2), ("Descending", 2)]
        third_columns_names = ["Total analysed",
                               "Semitones", "Interval", "Semitones", "Interval"]
        columns = columns_alike_our_data(
            third_columns_names, second_column_names)
        computations = ["sum", "max", "maxInterval", "min", "minInterval"]

        hoja_iValues(workbook.create_sheet("Largest_leaps"), columns, data, third_columns_names, computations,
                     sorting_lists, groups=groups, second_columns=second_column_names, average=True, additional_info=additional_info)

        if "Sheet" in workbook.get_sheet_names():  # Delete the default sheet
            std = workbook.get_sheet_by_name('Sheet')
            workbook.remove_sheet(std)
        workbook.save(os.path.join(results_path, name))

        with visualiser_lock:
            # VISUALISATIONS
            columns_visualisation = [
                'Intervallic ratio', 'Trimmed intervallic ratio', 'Std']
            if groups:
                data_grouped = data.groupby(list(groups))
                for g, datag in data_grouped:
                    result_visualisations = path.join(
                        results_path, 'visualisations', g)
                    if not os.path.exists(result_visualisations):
                        os.mkdir(result_visualisations)

                    name_bar = path.join(
                        result_visualisations, name.replace('.xlsx', '.png'))
                    ivalues_bar_plot(
                        name_bar, datag, columns_visualisation, second_title=str(g))
                    name_box = path.join(
                        result_visualisations, 'Ambitus' + name.replace('.xlsx', '.png'))
                    box_plot(name_box, datag, second_title=str(g))
            else:
                name_bar = results_path + \
                    path.join('visualisations', name.replace('.xlsx', '.png'))
                ivalues_bar_plot(name_bar, data, columns_visualisation)
                name_box = path.join(
                    results_path, 'visualisations', 'Ambitus' + name.replace('.xlsx', '.png'))
                box_plot(name_box, data)
    except Exception as e:
        logger.error('{}   Problem found:'.format(name), exc_info=True)


def iiaIntervals(data, name, sorting_list, results_path, sorting_lists, visualiser_lock, additional_info=[], groups=None):
    try:
        workbook = openpyxl.Workbook()
        all_columns = data.columns.tolist()
        general_cols = copy.deepcopy(not_used_cols)
        for row in rows_groups:
            if len(rows_groups[row][0]) == 0:
                general_cols.append(row)
            else:
                general_cols += rows_groups[row][0]

        # nombres de todos los intervalos
        third_columns_names_origin = set(all_columns) - set(general_cols)
        third_columns_names_origin = sort(
            third_columns_names_origin, sorting_list)
        third_columns_names = ['Total analysed'] + third_columns_names_origin

        # esta hoja va de sumar, así que en todas las columnas el cómputo que hay que hacer es sumar!
        computations = ["sum"]*len(third_columns_names)

        hoja_iValues(workbook.create_sheet("Weighted"), third_columns_names, data, third_columns_names, computations, sorting_lists,
                     groups=groups, average=True, last_column=True, last_column_average=False, additional_info=additional_info, ponderate=True)
        hoja_iValues(workbook.create_sheet("Horizontal Per"), third_columns_names, data, third_columns_names, computations, sorting_lists,
                     groups=groups, per=True, average=True, last_column=True, last_column_average=False, additional_info=additional_info)
        hoja_iValues(workbook.create_sheet("Vertical Per"), third_columns_names, data, third_columns_names, computations, sorting_lists,
                     groups=groups, per=True, average=False, last_column=True, last_column_average=True, additional_info=additional_info)

        if "Sheet" in workbook.get_sheet_names():  # Delete the default sheet
            std = workbook.get_sheet_by_name('Sheet')
            workbook.remove_sheet(std)
        workbook.save(os.path.join(results_path, name))

        with visualiser_lock:
            # VISUALISATIONS
            if 'Clefs' in name:
                title = 'Use of clefs'
            elif 'Intervals_absolute' in name:
                title = 'Presence of intervals (direction dismissed)'
            else:
                title = 'Presence of intervals (ascending and descending)'

            if groups:
                data_grouped = data.groupby(list(groups))
                for g, datag in data_grouped:
                    result_visualisations = path.join(
                        results_path, 'visualisations', g)
                    if not os.path.exists(result_visualisations):
                        os.mkdir(result_visualisations)

                    name_bar = path.join(
                        result_visualisations, name.replace('.xlsx', '.png'))
                    bar_plot(name_bar, datag, third_columns_names_origin,
                             'Intervals' if 'Clef' not in name else 'Clefs', title, second_title=str(g))
            else:
                name_bar = path.join(
                    results_path, 'visualisations', name.replace('.xlsx', '.png'))
                bar_plot(name_bar, data, third_columns_names_origin,
                         'Intervals' if 'Clef' not in name else 'Clefs', title)
    except Exception as e:
        logger.error('{}  Problem found:'.format(name), exc_info=True)

#########################################################
# Function to generate the file 3.Intervals_types.xlsx  #
#########################################################


def IIIIntervals_types(data, results_path, name, sorting_lists, visualiser_lock, groups=None, additional_info=[]):
    try:
        workbook = openpyxl.Workbook()

        second_column_names = [("", 2), ("Leaps", 3), ("StepwiseMotion", 3)]
        second_column_names2 = [('', 1), ("Perfect", 3), ("Major", 3),
                                ("Minor", 3), ("Augmented", 3), ("Diminished", 3)]
        third_columns_names = ['Total analysed', "RepeatedNotes",
                               "Ascending", "Descending", "All", "Ascending", "Descending", "All"]
        third_columns_names2 = ['Total analysed', "Ascending", "Descending", "All", "Ascending", "Descending", "All",
                                "Ascending", "Descending", "All", "Ascending", "Descending", "All", "Ascending", "Descending", "All"]
        computations = ["sum"]*len(third_columns_names)
        computations2 = ['sum']*len(third_columns_names2)
        columns = columns_alike_our_data(
            third_columns_names, second_column_names)
        columns2 = columns_alike_our_data(
            third_columns_names2, second_column_names2)

        hoja_iValues(workbook.create_sheet("Weighted"), columns, data, third_columns_names, computations, sorting_lists, groups=groups, last_column=True, last_column_average=False, second_columns=second_column_names, average=True,
                     columns2=columns2,  third_columns2=third_columns_names2, computations_columns2=computations2, second_columns2=second_column_names2, additional_info=additional_info, ponderate=True)
        hoja_iValues(workbook.create_sheet("Horizontal Per"), columns, data, third_columns_names, computations, sorting_lists, groups=groups, second_columns=second_column_names, per=True, average=True, last_column=True, last_column_average=False,
                     columns2=columns2,  third_columns2=third_columns_names2, computations_columns2=computations2, second_columns2=second_column_names2, additional_info=additional_info)
        hoja_iValues(workbook.create_sheet("Vertical Per"), columns, data, third_columns_names, computations, sorting_lists, groups=groups, second_columns=second_column_names, per=True, average=False, last_column=True, last_column_average=True,
                     columns2=columns2,  third_columns2=third_columns_names2, computations_columns2=computations2, second_columns2=second_column_names2, additional_info=additional_info)

        # borramos la hoja por defecto
        if "Sheet" in workbook.get_sheet_names():
            std = workbook.get_sheet_by_name('Sheet')
            workbook.remove_sheet(std)
        workbook.save(os.path.join(results_path, name))

        with visualiser_lock:
            # VISUALISATIONS
            if groups:
                data_grouped = data.groupby(list(groups))
                for g, datag in data_grouped:
                    result_visualisations = path.join(
                        results_path, 'visualisations', g)
                    if not os.path.exists(result_visualisations):
                        os.mkdir(result_visualisations)

                    name1 = path.join(result_visualisations,
                                      name.replace('.xlsx',  '') + '_AD.png')
                    pie_plot(name1, datag, second_title=str(g))
                    name2 = path.join(result_visualisations,
                                      name.replace('.xlsx',  '.png'))
                    double_bar_plot(name2, data, second_title=str(g))
            else:
                name1 = path.join(results_path, 'visualisations',
                                  name.replace('.xlsx', '') + '_AD.png')
                pie_plot(name1, data)
                name2 = path.join(results_path, 'visualisations',
                                  name.replace('.xlsx',  '.png'))
                double_bar_plot(name2, data)
    except Exception as e:
        logger.error('3Interval_types  Problem found:', exc_info=True)
#######################################################################################
# Function used to generate the files 2a.Intervals, 2a.Intervals_absolute and 5.Clefs #
#######################################################################################

###########################################################################################################################################################
# This function is in charge of printing each iValue's sheet
#
#   hoja: the openpyxl sheet object in which we will write
#   columns: list of the dataframe (grouped) column names that we need to access (as it doesn't necessarily correspond to the names that we want to print)
#   data: main dataframe
#   third_columns: list of the names of the columns that we need to print
#   computations_columns: information about the matematical computation that has to be done to each column (sum, mean...)
#   sorting_lists: dictionary of lists used for sorting the output and showing it in an appropiate way
#   ----------------
#   groups: used for factor > 1
#   first_columns: list of column names to print in first place, along with the number of columns that each has to embrace
#   second_columns: list of column names to print in second place, along with the number of columns that each has to embrace
#   per: boolean value to indicate if we need to compute the excel in absolute values or percentage (by default it is absolute)
#   average: boolean value to indicate if we want the average row at the last group's row
#   last_column: boolean value to indicate if we want a summarize on the last column
#   last_column_average: boolean to indicate if we want the last column to have each row's average (otherwise, the total is writen)
#   ------
#   columns2: names of the second groups of columns (some sheets have subgroupings at the right), used in emphatised_scale_degrees and Intervals_types
#   data2: dataframe used for printing information at the right
#   third_columns2: columns that will be printed
#   computations_columns2: computations for those columns
#   first_columns2: columns printed on first place
#   second_columns2: colummns printed on second place
#   additional_info: list of additional columns
#   ponderate: boolean to indicate if we want to ponderate the data printed or not
#
###########################################################################################################################################################
