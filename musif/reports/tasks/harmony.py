import copy
import os
from multiprocessing import Lock
from typing import Dict, List

import openpyxl
import pandas as pd
from musif.common.sort import sort_columns, sort_list
from musif.config import Configuration
from musif.extract.features.harmony.constants import (
    HARMONIC_RHYTHM,
    HARMONIC_RHYTHM_BEATS,
    KEY_MODULATORY,
    KEY_PERCENTAGE,
    KEY_PREFIX,
    ADDITIONS_prefix,
    CHORD_prefix,
    CHORD_TYPES_prefix,
    CHORDS_GROUPING_prefix,
    NUMERALS_prefix,
)
from musif.logs import pwarn
from musif.reports.constants import (
    COMMON_DF,
    EXCEPTIONS,
    IMAGE_EXTENSION,
    NARROW,
    NORMAL_WIDTH,
    TOTAL_ANALYSED,
    VISUALIZATIONS,
    not_used_cols,
)
from musif.reports.tasks.sort_labels import sort_labels
from musif.reports.utils import (
    create_excel,
    get_excel_name,
    get_general_cols,
    print_basic_sheet,
    save_workbook,
)
from musif.reports.visualizations import (
    bar_plot_extended,
    bar_plot_percentage,
    line_plot_extended,
)
from pandas.core.frame import DataFrame

rows_groups_harmony = {}
visualizations_harmony = bool
visualizations_folder = ""


def Harmonic_analysis(
    rows_groups: dict,
    not_used_cols: dict,
    factor,
    _cfg: Configuration,
    kwargs: DataFrame,
    pre_string,
    name: str,
    results_path: str,
    visualizations: Lock,
    additional_info: list = [],
    groups: list = None,
):
    try:
        workbook = openpyxl.Workbook()
        excel_name = get_excel_name(pre_string, name)

        general_cols = copy.deepcopy(not_used_cols)
        get_general_cols(rows_groups, general_cols)

        global rows_groups_harmony
        rows_groups_harmony = rows_groups
        global visualizations_harmony
        visualizations_harmony = visualizations
        global visualizations_folder
        visualizations_folder = os.path.join(results_path, VISUALIZATIONS)

        Print_Harmonic_Data(factor, _cfg, kwargs, additional_info, groups, workbook)
        Print_Chords(factor, _cfg, kwargs, groups, additional_info, workbook)
        Print_Functions(factor, _cfg, kwargs, groups, additional_info, workbook)
        Print_Keys(factor, _cfg, kwargs, groups, additional_info, workbook)
        save_workbook(
            os.path.join(results_path, excel_name), workbook, cells_size=NARROW
        )

    except Exception as e:
        pwarn("{}  Problem found: {}".format(name, e))


def Print_Harmonic_Data(
    factor: int,
    _cfg: Configuration,
    info: DataFrame,
    additional_info: DataFrame,
    groups: Dict,
    workbook: openpyxl.Workbook,
) -> None:
    data = info["harmonic_data"]
    harmonic_rythm = [c for c in data.columns if "Harmonic_rhythm" in c]

    chordTypes = [c for c in data.columns if CHORD_TYPES_prefix in c]
    chordTypes = sort_list(chordTypes, _cfg.sorting_lists["ChordTypeGrouppingSorting"])

    additions = [c for c in data.columns if ADDITIONS_prefix in c]
    second_column = [
        ("", 1),
        ("Harmonic Rhythm", len(harmonic_rythm)),
        ("Chord types", len(chordTypes)),
        ("Additions", len(additions)),
    ]
    data = data[harmonic_rythm + chordTypes + additions]

    data.columns = [
        i.replace(CHORD_TYPES_prefix, "")
        .replace(NUMERALS_prefix, "")
        .replace(ADDITIONS_prefix, "")
        for i in data
    ]
    data.rename(
        columns={HARMONIC_RHYTHM: "Bars", HARMONIC_RHYTHM_BEATS: "Beats"}, inplace=True
    )
    third_columns = ["Total analysed"] + list(data.columns)

    data_total = pd.concat((info[COMMON_DF], data), axis=1)
    print_basic_sheet(
        _cfg,
        rows_groups_harmony,
        "Harmonic Data",
        data_total,
        additional_info,
        groups,
        workbook,
        second_column,
        third_columns,
    )

    if visualizations_harmony:
        create_visualizations(
            factor,
            data_total,
            ["Bars", "Beats"],
            groups,
            "Harmonic Rhythm",
            type="extended",
        )
        create_visualizations(
            factor,
            data_total,
            [i.replace(CHORD_TYPES_prefix, "") for i in chordTypes],
            groups,
            "Chord Types",
        )
        create_visualizations(
            factor,
            data_total,
            [i.replace(ADDITIONS_prefix, "") for i in additions],
            groups,
            "Additions",
        )


def Print_Chords(
    factor: int,
    _cfg: Configuration,
    info: dict,
    groups: dict,
    additional_info: DataFrame,
    workbook: openpyxl.Workbook,
) -> None:

    data = info["chords"]
    data.columns = [i.replace(CHORDS_GROUPING_prefix, "") for i in data.columns]
    data.columns = [i.replace(CHORD_prefix, "") for i in data.columns]
    data.columns = [i.replace("_Count", "") for i in data.columns]

    try:
        data = sort_columns(
            data,
            sort_labels(
                data.columns, chordtype="occurrences", form=["", "+", "o", "%", "M"]
            ),
        )
    except:
        data = sort_columns(data, _cfg.sorting_lists["NumeralsSorting"])

    third_columns = data.columns.to_list()

    third_columns.insert(0, "Total analysed")
    data = pd.concat([info[COMMON_DF], data], axis=1)

    computations = ["sum"] + ["mean"] * (len(third_columns) - 1)

    create_excel(
        workbook.create_sheet("Chords Weighted"),
        rows_groups_harmony,
        third_columns,
        data,
        third_columns,
        computations,
        _cfg.sorting_lists,
        groups=groups,
        per=True,
        average=True,
        last_column=True,
        last_column_average=False,
        additional_info=additional_info,
    )
    if visualizations_harmony:
        create_visualizations(factor, data, third_columns, groups, "Chords")


def Print_Keys(
    factor: int,
    _cfg: Configuration,
    info: dict,
    groups: dict,
    additional_info: DataFrame,
    workbook: openpyxl.Workbook,
) -> None:
    data = info["key_areas"]
    data1 = data[[i for i in data.columns if KEY_MODULATORY in i]]
    data2 = data[[i for i in data.columns if KEY_PERCENTAGE in i]]

    data1.columns = [
        i.replace(KEY_MODULATORY, "").replace("Key", "").replace("_", " ")
        for i in data1.columns
    ]
    data2.columns = [
        i.replace(KEY_PERCENTAGE, "").replace("Key", "").replace("_", " ")
        for i in data2.columns
    ]

    third_columns = ["Total analysed"] + data1.columns.tolist()
    third_columns2 = ["Total analysed"] + data2.columns.tolist()

    second_columns = [("", 1), ("No.", len(third_columns))]
    second_columns2 = [("", 1), ("Measures", len(third_columns2))]

    Print_Double_Excel(
        "Key Areas",
        factor,
        _cfg,
        groups,
        additional_info,
        workbook,
        info[COMMON_DF],
        data1,
        data2,
        third_columns,
        third_columns2,
        second_column_names=second_columns,
        second_column_names2=second_columns2,
    )

    if visualizations_harmony:
        columns = [i for i in data.columns if KEY_MODULATORY in i]
        columns2 = [i for i in data.columns if KEY_PERCENTAGE in i]
        columns = columns + columns2

        data_total = pd.concat((info["common_df"], data), axis=1)

        create_visualizations(factor, data_total, columns, groups, "Keys")


def Print_Functions(
    factor: int,
    _cfg: Configuration,
    info: DataFrame,
    groups: dict,
    additional_info: DataFrame,
    workbook: openpyxl.Workbook,
) -> None:
    data = info["functions"]
    (
        data1,
        data2,
        data3,
        third_columns_names,
        third_columns_names2,
        third_columns_names3,
        second_column_names,
        second_column_names2,
        second_column_names3,
    ) = Process_data(
        data,
        NUMERALS_prefix,
        CHORDS_GROUPING_prefix,
        CHORDS_GROUPING_prefix,
        _cfg.sorting_lists["NumeralsSorting"],
        _cfg.sorting_lists["ModulationG2Sorting"],
        _cfg.sorting_lists["ModulationG2Sorting"],
    )

    Print_Triple_Excel(
        "Functions",
        factor,
        _cfg,
        groups,
        additional_info,
        workbook,
        info[COMMON_DF],
        data1,
        data2,
        data3,
        third_columns_names,
        third_columns_names2,
        third_columns_names3,
        second_column_names,
        second_column_names2,
        second_column_names3,
    )

    if visualizations_harmony:
        data.columns = [i.replace("Numerals_", "") for i in data.columns]
        data_total = pd.concat((info["common_df"], data), axis=1)
        create_visualizations(
            factor, data_total, third_columns_names, groups, "Functions"
        )


def Print_Triple_Excel(
    name: str,
    factor: int,
    _cfg: Configuration,
    groups: dict,
    additional_info: DataFrame,
    workbook: openpyxl.Workbook,
    data_general: DataFrame,
    data1: DataFrame,
    data2: DataFrame,
    data3: DataFrame,
    third_columns_names: List[str],
    third_columns_names2: List[str],
    third_columns_names3: List[str],
    second_column_names: List[str],
    second_column_names2: List[str],
    second_column_names3: List[str],
) -> None:
    computations = ["sum"] + ["mean"] * (len(third_columns_names))
    computations2 = ["sum"] + ["mean"] * (len(third_columns_names2))
    computations3 = ["sum"] + ["mean"] * (len(third_columns_names3))

    data1 = pd.concat([data_general, data1], axis=1)
    data2 = pd.concat([data_general, data2], axis=1)
    data3 = pd.concat([data_general, data3], axis=1)

    create_excel(
        workbook.create_sheet(name + " Weighted"),
        rows_groups_harmony,
        third_columns_names,
        data1,
        third_columns_names,
        computations,
        _cfg.sorting_lists,
        groups=groups,
        data2=data2,
        second_columns=second_column_names,
        columns2=third_columns_names2,
        third_columns2=third_columns_names2,
        computations_columns2=computations2,
        second_columns2=second_column_names2,
        data3=data3,
        columns3=third_columns_names3,
        third_columns3=third_columns_names3,
        computations_columns3=computations3,
        second_columns3=second_column_names3,
        additional_info=additional_info,
        per=True,
        average=True,
        last_column=True,
        last_column_average=False,
    )


def Print_Double_Excel(
    name: str,
    factor: int,
    _cfg: Configuration,
    groups: dict,
    additional_info: dict,
    workbook: openpyxl.Workbook,
    data_general: DataFrame,
    data1: DataFrame,
    data2: DataFrame,
    third_columns_names: list,
    third_columns_names2: list,
    second_column_names: List[str] = None,
    second_column_names2: List[str] = None,
) -> None:
    computations = ["sum"] + ["mean"] * (len(third_columns_names))
    computations2 = ["sum"] + ["mean"] * (len(third_columns_names2))

    data1 = pd.concat([data_general, data1], axis=1)
    data2 = pd.concat([data_general, data2], axis=1)

    create_excel(
        workbook.create_sheet(name),
        rows_groups_harmony,
        third_columns_names,
        data1,
        third_columns_names,
        computations,
        _cfg.sorting_lists,
        groups=groups,
        second_columns=second_column_names,
        data2=data2,
        columns2=third_columns_names2,
        third_columns2=third_columns_names2,
        computations_columns2=computations2,
        second_columns2=second_column_names2,
        additional_info=additional_info,
        per=True,
        average=True,
        last_column=True,
        last_column_average=False,
    )


def Process_data(
    data: DataFrame,
    prefix1: str,
    prefix2: str,
    prefix3: str,
    sorting_list1: list,
    sorting_list2: list,
    sorting_list3: list,
) -> None:
    data.columns = [i.replace("_Count", "") for i in data.columns]
    data1 = data[[i for i in data.columns if prefix1 in i]]
    data1.columns = [i.replace(prefix1, "").replace("_", " ") for i in data1.columns]

    data2 = data[[i for i in data.columns if prefix2 + "1" in i]]
    data2.columns = [
        i.replace(prefix2 + "1", "").replace("_", " ") for i in data2.columns
    ]

    data3 = data[[i for i in data.columns if prefix2 + "2" in i]]
    data3.columns = [
        i.replace(prefix3 + "2", "Grouped").replace("_", " ") for i in data3.columns
    ]

    data1 = sort_columns(data1, sorting_list1)
    data2 = sort_columns(data2, sorting_list2)
    data3 = sort_columns(data3, sorting_list3)

    data2 = data2.round(decimals=2)
    data3 = data3.round(decimals=2)

    third_columns_names = list(data1.columns)
    third_columns_names2 = list(data2.columns)
    third_columns_names3 = list(data3.columns)

    insert_total_analysed(third_columns_names)
    insert_total_analysed(third_columns_names2)
    insert_total_analysed(third_columns_names3)

    second_column_names = name_second_columns(third_columns_names, prefix1)
    second_column_names2 = name_second_columns(third_columns_names2, prefix2)
    second_column_names3 = name_second_columns(third_columns_names3, prefix3 + " 2")

    return (
        data1,
        data2,
        data3,
        third_columns_names,
        third_columns_names2,
        third_columns_names3,
        second_column_names,
        second_column_names2,
        second_column_names3,
    )


def create_visualizations(
    factor: int,
    data: DataFrame,
    columns: list,
    groups: dict,
    name: str,
    type: str = "percentage",
) -> None:
    if TOTAL_ANALYSED in columns:
        columns.remove(TOTAL_ANALYSED)
    title = name.capitalize().replace(" ", "_")

    if groups:
        data_grouped = data.groupby(list(groups))
        for g, datag in data_grouped:
            visualisations_path = name + VISUALIZATIONS + str(g.replace("/", "_"))
            if not os.path.exists(visualisations_path):
                os.mkdir(visualisations_path)
            name_bar = (
                visualisations_path + "\\" + name.replace(".xlsx", IMAGE_EXTENSION)
            )
            bar_plot_extended(
                name_bar,
                datag,
                columns,
                title,
                "Percentage",
                title,
                second_title=str(g),
            )

    elif factor == 1:
        groups = [i for i in rows_groups_harmony]
        for row in rows_groups_harmony:
            plot_name = (
                name.replace(".xlsx", "") + "_Per_" + str(row.upper()) + IMAGE_EXTENSION
            )

            name_bar = os.path.join(visualizations_folder, plot_name)

            if row not in not_used_cols:
                if len(rows_groups_harmony[row][0]) == 0:  # no sub-groups
                    data_grouped = data.groupby(row, sort=True)
                    if data_grouped:
                        line_plot_extended(
                            name_bar,
                            data_grouped,
                            columns,
                            "Instrument",
                            "Density",
                            title,
                            second_title="Per " + str(row),
                        )
                else:
                    for i, subrow in enumerate(rows_groups_harmony[row][0]):
                        if subrow not in EXCEPTIONS:
                            plot_name = (
                                name.replace(".xlsx", "")
                                + "_Per_"
                                + str(row.upper())
                                + "_"
                                + str(subrow)
                                + IMAGE_EXTENSION
                            )
                            name_bar = os.path.join(visualizations_folder, plot_name)
                            data_grouped = data.groupby(subrow)
                            line_plot_extended(
                                name_bar,
                                data_grouped,
                                columns,
                                "Instrument",
                                "Density",
                                title,
                                second_title="Per " + str(subrow),
                            )
    else:
        name_bar = os.path.join(visualizations_folder, name + IMAGE_EXTENSION)
        if type == "extended":
            bar_plot_extended(name_bar, data, columns, name, "Value", title)
        elif type == "percentage":
            bar_plot_percentage(name_bar, data, columns, name, title)


def name_second_columns(third_columns_names: List[str], name: str) -> List[str]:
    second_column_names = [("", 1), (name.replace("_", ""), len(third_columns_names))]
    return second_column_names


def insert_total_analysed(columns_names: List[str]) -> None:
    columns_names.insert(0, "Total analysed")
