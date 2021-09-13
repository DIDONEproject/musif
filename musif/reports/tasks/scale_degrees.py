import copy
from musif.reports.visualisations import customized_plot
import os
from multiprocessing import Lock
from os import path
import numpy as np

import pandas as pd
from musif.common.sort import sort

import musif.extract.features.ambitus as ambitus
import musif.extract.features.lyrics as lyrics
import openpyxl
from config import Configuration
from music21 import interval
from musif.reports.constants import *
from musif.reports.utils import excel_sheet
from musif.reports.utils import adjust_excel_width_height, prepare_data_emphasised_scale_degrees_second
from pandas.core.frame import DataFrame

########################################################################
# Function to generate the reports files Emphasised_scale_degrees.xlsx
########################################################################

def Emphasised_scale_degrees(rows_groups, not_used_cols, factor, _cfg: Configuration, data: DataFrame, sorting_list: list, name: str, results_path: str, visualiser_lock: Lock, groups: list=None, additional_info=[]):
    try:
        workbook = openpyxl.Workbook()
        all_columns = data.columns.tolist()
        general_cols = copy.deepcopy(not_used_cols)
        for row in rows_groups:
            if len(rows_groups[row][0]) == 0:
                general_cols.append(row)
            else:
                general_cols += rows_groups[row][0]

        # nombres de todos los intervalos
        third_columns_names_origin = list(set(all_columns) - set(general_cols))
        third_columns_names_origin = sort(
            third_columns_names_origin, sorting_list)
        third_columns_names = ['Total analysed'] + third_columns_names_origin
        third_columns_names2 = ['Total analysed'] + \
            ['1', '4', '5', '7', 'Others']

        # esta sheet va de sumar, así que en todas las columnas el cómputo que hay que hacer es sumar!
        computations = ["sum"]*len(third_columns_names)
        computations2 = ["sum"]*len(third_columns_names2)

        emph_degrees = pd.DataFrame(prepare_data_emphasised_scale_degrees_second(
            data, third_columns_names, third_columns_names2))
        data2 = pd.concat(
            [data[[gc for gc in general_cols if gc in all_columns]], emph_degrees], axis=1)
        _, unique_columns = np.unique(data2.columns, return_index=True)
        data2 = data2.iloc[:, unique_columns]

        excel_sheet(workbook.create_sheet("Weighted"), third_columns_names, data, third_columns_names, computations, _cfg.sorting_lists, groups=groups, last_column=True, last_column_average=False, average=True,
                     columns2=third_columns_names2,  data2=data2, third_columns2=third_columns_names2, computations_columns2=computations2, additional_info=additional_info, ponderate=True)
        if factor>=1:
            excel_sheet(workbook.create_sheet("Horizontal Per"), third_columns_names, data, third_columns_names, computations, _cfg.sorting_lists, groups=groups, per=True, average=True, last_column=True, last_column_average=False,
                     columns2=third_columns_names2,  data2=data2, third_columns2=third_columns_names2, computations_columns2=computations2, additional_info=additional_info)

        # Delete the default sheet
        if "Sheet" in workbook.get_sheet_names():
            std = workbook.get_sheet_by_name('Sheet')
            workbook.remove_sheet(std)
        adjust_excel_width_height(workbook)
        workbook.save(os.path.join(results_path, name))

        # with visualiser_lock:
        subtitile = 'in relation to the global key' if '4a' in name else 'in relation to the local key'
            # VISUALISATIONS
        if groups:
            data_grouped = data.groupby(list(groups))
            for g, datag in data_grouped:
                result_visualisations = path.join(
                    results_path, 'visualisations', g)
                if not os.path.exists(result_visualisations):
                    os.mkdir(result_visualisations)

                name1 = path.join(
                    result_visualisations, 'Scale_degrees_GlobalKey.png' if 'A' in name else 'Scale_degrees_LocalKey.png')
                customized_plot(
                    name1, data, third_columns_names_origin, subtitile, second_title=g)
        else:
            name1 = path.join(results_path, 'visualisations',
                                'Scale_degrees_GlobalKey.png' if 'A' in name else 'Scale_degrees_LocalKey.png')
            customized_plot(
                name1, data, third_columns_names_origin, subtitile)

    except Exception as e:
        _cfg.write_logger.info('{}  Problem found: {}'.format(name, e))

