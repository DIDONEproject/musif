import os
from multiprocessing import Lock
from os import path
from typing import Dict

import numpy as np
import openpyxl
from music21 import interval
import pandas as pd
from pandas.core.frame import DataFrame
from musif.extract.features.ambitus.constants import HIGHEST_NOTE, HIGHEST_NOTE_INDEX, LOWEST_NOTE, LOWEST_NOTE_INDEX
from musif.extract.features.interval.constants import ABSOLUTE_INTERVALLIC_STD, ABSOLUTE_INTERVALLIC_TRIM_DIFF, ABSOLUTE_INTERVALLIC_TRIM_RATIO, ASCENDING_INTERVALLIC_MEAN, DESCENDING_INTERVALLIC_MEAN, INTERVALLIC_STD, LARGEST_INTERVAL_ASC, LARGEST_INTERVAL_DESC, TRIMMED_ABSOLUTE_INTERVALLIC_MEAN

import musif.extract.features.lyrics as lyrics
from musif.config import Configuration
from musif.extract.features import ambitus, interval as I
from musif.logs import pwarn
from musif.reports.constants import CLEF1, EXCEPTIONS, IMAGE_EXTENSION, NARROW, ROLE, COMMON_DF, VISUALIZATIONS
from musif.reports.utils import create_excel, columns_alike_our_data, get_excel_name, save_workbook
from musif.reports.visualizations import box_plot, melody_bar_plot

INTERVALLIC_MEAN = "Intervallic Mean"
TRIMMED_INTERVALLIC_MEAN = 'Trimmed Intervallic Mean'
DIFF_TRIMMED = "dif. Trimmed"
STD = "Std"
ABSOLUTE_INTERVALLIC_MEAN = 'Absolute Intervallic Mean'
ABSOLUTE_STD = "Absolute Std"
TRIM_RATIO = "% Trimmed"

def Melody_values(rows_groups, not_used_cols, factor, _cfg: Configuration, info: Dict[str, DataFrame], results_path: str, pre_string, name: str, visualizations: Lock, additional_info: list=[], remove_columns: bool=False, groups: list=None):
    try:
        excel_name = get_excel_name(pre_string, name)
        workbook = openpyxl.Workbook()
        data=info['melody_values']
        Rename_columns(data)

        # data_general = data[metadata_columns+ ['Total analysed']]

        data_general = info[COMMON_DF]
        data_general['Total analysed']=1.0
        data = pd.concat([data_general,  data], axis=1)
        
        PrintStatisticalValues(_cfg, data, additional_info, groups, workbook, rows_groups)
        PrintAmbitus(_cfg, data, additional_info, remove_columns, groups, workbook,rows_groups)
        PrintLargestLeaps(_cfg, data, additional_info, groups, workbook, rows_groups)
        save_workbook(os.path.join(results_path, excel_name), workbook, cells_size=NARROW)

        if visualizations:
            columns_visualizations = [INTERVALLIC_MEAN, TRIMMED_INTERVALLIC_MEAN, STD, ABSOLUTE_INTERVALLIC_MEAN,ABSOLUTE_STD]
            path_visualizations = path.join(results_path,VISUALIZATIONS)

            if not os.path.exists(path_visualizations):
                os.makedirs(path_visualizations)
            if groups:
                data_grouped = data.groupby(list(groups))
                for g, datag in data_grouped:
                    path_visualizations = os.path.join(
                        path_visualizations, g)
                    if not os.path.exists(path_visualizations):
                        os.mkdir(path_visualizations)

                    name_bar = os.path.join(
                        path_visualizations, name.replace('.xlsx', IMAGE_EXTENSION))
                    melody_bar_plot(
                        name_bar, datag, columns_visualizations, second_title=str(g))
                    name_box = path.join(
                        path_visualizations, 'Ambitus' + name.replace('.xlsx', IMAGE_EXTENSION))
                    box_plot(name_box, datag, second_title=str(g))

            elif factor == 1:
                for row in rows_groups:
                    plot_name = name.replace(
                        '.xlsx', '') + '_Per_' + str(row.replace('Aria','').upper()) + IMAGE_EXTENSION
                    name_bar =path.join(path_visualizations,'Per_'+row.replace('Aria','').upper())
                    if not os.path.exists(name_bar):
                        os.makedirs(name_bar)
                    name_bar=path.join(name_bar,plot_name)
                    if row not in not_used_cols:
                        if len(rows_groups[row][0]) == 0:  # no sub-groups
                            data_grouped = data.groupby(row, sort=True)
                            if data_grouped:
                                melody_bar_plot(name_bar, data_grouped, columns_visualizations, second_title='Per ' + str(row.replace('Aria','').upper()))
                                if row == CLEF1: #Exception for boxplots
                                    name_box = name_bar.replace('Melody_Values', 'Ambitus')
                                    box_plot(name_box, data_grouped, second_title='Per '+ str(row.replace('Aria','').upper()))
                        else: #subgroups
                                for i, subrow in enumerate(rows_groups[row][0]):
                                    if subrow not in EXCEPTIONS:
                                        name_bar=name_bar.replace(IMAGE_EXTENSION,'')+'_'+subrow+IMAGE_EXTENSION
                                        data_grouped = data.groupby(subrow)
                                        melody_bar_plot(name_bar, data_grouped, columns_visualizations, second_title='Per ' + str(subrow.replace('Aria','').upper()))
                                        name_box = path.join(
                                        path_visualizations, 'Ambitus' + name.replace('.xlsx', IMAGE_EXTENSION))
                                        
                                        if subrow == ROLE:
                                            box_plot(name_box, data_grouped, second_title='Per '+ str(subrow.replace('Aria','').upper()))
            else:                   
                name_bar = path.join(path_visualizations, name.replace('.xlsx', IMAGE_EXTENSION))
                melody_bar_plot(name_bar, data, columns_visualizations)
                name_box = path.join(path_visualizations, 'Ambitus' + name.replace('.xlsx', IMAGE_EXTENSION))
                box_plot(name_box, data)
    except Exception as e:
        pwarn('{}  Problem found: {}'.format(name, e))


def PrintLargestLeaps(_cfg, data, additional_info, groups, workbook, rows_groups):
    second_column_names = [("", 1), ("Ascending", 2), ("Descending", 2)]
    third_columns_names = ["Total analysed",
                               "Semitones", "Interval", "Semitones", "Interval"]
    columns = columns_alike_our_data(
            third_columns_names, second_column_names)

    computations = ["sum", "max", "maxInterval", "min", "minInterval"]

    create_excel(workbook.create_sheet("Largest_leaps"), rows_groups, columns, data, third_columns_names, computations,
                     _cfg.sorting_lists, groups=groups, second_columns=second_column_names, average=True, additional_info=additional_info)

def PrintAmbitus(_cfg, data, additional_info, remove_columns, groups, workbook, rows_groups):
    first_column_names = [("", 1), ("Lowest", 2), ("Highest", 2), ("Lowest", 2), ("Highest", 2), (
            "Ambitus", 6)] if not remove_columns else [("", 1), ("Lowest", 2), ("Highest", 2), ("Ambitus", 2)]

    second_column_names = [("", 5), ("Mean", 2), ("Mean", 2), ("Largest", 2), ("Smallest", 2), ("Mean", 2)] if not remove_columns else [("", 5), ("Largest", 2)]

    third_columns_names = ["Total analysed", "Note", "Index", "Note", "Index", "Note", "Index", "Note", "Index", "Semitones", "Interval", "Semitones", "Interval",
         "Semitones", "Interval"] if not remove_columns else ["Total analysed", "Note", "Index", "Note", "Index", "Semitones", "Interval"]

    computations = ["sum", "minNote", "min", "maxNote", "max", 'meanNote', 'mean', 'meanNote', 'mean', 'max', "maxInterval", 'min', "minInterval", "mean", "meanInterval"] if not remove_columns else ["sum", "minNote", "min", "maxNote", "max", 'max', "maxInterval"]

    columns = columns_alike_our_data(
            third_columns_names, second_column_names, first_column_names)
    columns = [i.replace('Ambitus', '') for i in columns]


    create_excel(workbook.create_sheet("Ambitus"), rows_groups, columns, data, third_columns_names, computations, _cfg.sorting_lists, groups=groups,
                     first_columns=first_column_names, second_columns=second_column_names, average=True, additional_info=additional_info)

def PrintStatisticalValues(_cfg, data, additional_info, groups, workbook, rows_groups):
    column_names = ["Total analysed", INTERVALLIC_MEAN, TRIMMED_INTERVALLIC_MEAN, DIFF_TRIMMED,
                       TRIM_RATIO, ABSOLUTE_INTERVALLIC_MEAN, STD, ABSOLUTE_STD]
    data.rename(columns={'IntervallicMean':INTERVALLIC_MEAN,'AbsoluteIntervallicMean':ABSOLUTE_INTERVALLIC_MEAN}, inplace=True)

    if lyrics.SYLLABIC_RATIO in data.columns:
        data.rename(columns={lyrics.SYLLABIC_RATIO: 'Syllabic ratio'}, inplace=True)
        
        column_names.append('Syllabic ratio')

    computations = ['sum'] + ["mean"]*(len(column_names) - 1)
    create_excel(workbook.create_sheet("Statistical_values"), rows_groups, column_names, data, column_names, computations,
                    _cfg.sorting_lists, groups=groups, average=True, additional_info=additional_info, ponderate=True)

def Rename_columns(data):
    data['LowestMeanIndex']=data[LOWEST_NOTE_INDEX]
    data['LowestMeanNote']=data[LOWEST_NOTE]
    data['HighestMeanIndex']=data[HIGHEST_NOTE_INDEX]
    data['HighestMeanNote']=data[HIGHEST_NOTE]

    data['MeanSemitones']= [interval.Interval(i).semitones if str(i) != 'nan' else np.nan for i in data['MeanInterval']]
    data.rename(columns={ambitus.LOWEST_NOTE_INDEX: "LowestIndex", HIGHEST_NOTE_INDEX: "HighestIndex"}, inplace=True)
    data.rename(columns={INTERVALLIC_MEAN: INTERVALLIC_MEAN, TRIMMED_ABSOLUTE_INTERVALLIC_MEAN: TRIMMED_INTERVALLIC_MEAN, ABSOLUTE_INTERVALLIC_TRIM_DIFF: DIFF_TRIMMED,
                             ABSOLUTE_INTERVALLIC_MEAN: ABSOLUTE_INTERVALLIC_MEAN,INTERVALLIC_STD: STD,ABSOLUTE_INTERVALLIC_STD: ABSOLUTE_STD, ABSOLUTE_INTERVALLIC_TRIM_RATIO:TRIM_RATIO}, inplace=True)

    data.rename(columns={LARGEST_INTERVAL_ASC: "AscendingInterval",ASCENDING_INTERVALLIC_MEAN: "AscendingSemitones", 
    LARGEST_INTERVAL_DESC: "DescendingInterval", DESCENDING_INTERVALLIC_MEAN: "DescendingSemitones"}, inplace=True)
    
    data.columns=[i.replace('All', '').replace('_','') for i in data.columns]
