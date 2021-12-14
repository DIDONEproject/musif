import copy
import os
from typing import Dict, List

import numpy as np
from musif.common.constants import VOICE_FAMILY
from musif.common.sort import sort_dataframe
from musif.reports.calculations import compute_average
from musif.reports.constants import (ARIA_ID, BOLD, CHARACTER, FONT,
                                     FONT_TITLE, GENDER, NAME, NORMAL_WIDTH,
                                     ROLE, YELLOWFILL, alfa, center,
                                     factors_Fill, fills_list, forbiden_groups,
                                     greenFill, interval, orangeFill, titles1Fill, titles2Fill,
                                     titles3Fill)
from openpyxl.utils import get_column_letter
from openpyxl.writer.excel import ExcelWriter
from pandas.core.frame import DataFrame

from .calculations import compute_value

WIDTH = 35
HEIGHT = 20

def Create_excel(sheet: ExcelWriter, rows_groups: Dict[str, list], columns: list, data: DataFrame, third_columns: list, computations_columns: list, sorting_lists: list, groups: list=None, first_columns: list=None,
                second_columns: list=None, per: bool=False, average: bool=False, last_column: bool=False, last_column_average: bool=False,
                columns2: list=None, data2: DataFrame=None, third_columns2: list=None, computations_columns2: list=None, first_columns2: list=None, second_columns2: list=None, 
                columns3: list=None, data3: DataFrame=None, third_columns3: list=None, computations_columns3: list=None, first_columns3: list=None, second_columns3: list=None, 
                additional_info: list=[], ponderate: bool=False):
    
    row_number = 1
    column_number = 1

    if groups == None:
        row_iteration(sheet, rows_groups, columns, row_number, column_number, data, third_columns, computations_columns, sorting_lists, first_columns=first_columns, second_columns=second_columns, per=per,
                    average=average, last_column=last_column, last_column_average=last_column_average,
                    columns2=columns2, data2=data2, third_columns2=third_columns2, computations_columns2=computations_columns2, first_columns2=first_columns2, second_columns2=second_columns2, 
                    columns3=columns3, data3=data3, third_columns3=third_columns3, computations_columns3=computations_columns3, first_columns3=first_columns3, second_columns3=second_columns3, 
                    additional_info=additional_info, ponderate=ponderate)
    else: # we may be grouping by more than 2 factors
        data_grouped = data.groupby(list(groups))

        last_printed = {i: ('', 0) for i in range(len(groups))}
        for group, group_data in data_grouped:
            cnumber = column_number
            group = [group] if type(group) != tuple else group
            for i, g in enumerate(group):
                if last_printed[i][0] != g:
                    sheet.cell(row_number, cnumber).value = g
                    sheet.cell(row_number, cnumber).fill = factors_Fill[i]
                    sheet.cell(row_number, column_number).font =  BOLD
                    counter_g = data[groups[i]].tolist().count(g)
                    sheet.cell(row_number, cnumber + 1).value = counter_g
                    if last_printed[i][0] != '':
                        sheet.merge_cells(
                            start_row=last_printed[i][1], start_column=i + 1, end_row=row_number - 2, end_column=i + 1)
                        sheet.cell(last_printed[i][1],
                                  i + 1).fill = factors_Fill[i]

                    last_printed[i] = (g, row_number + 1)

                row_number += 1
                cnumber += 1
            data2_grouped = None
            if data2 is not None:
                data2_grouped = data2.groupby(list(groups)).get_group(
                    group if len(group) > 1 else group[0])
            rn = row_iteration(sheet,rows_groups, columns, row_number, cnumber, group_data, third_columns, computations_columns, sorting_lists, group=groups, first_columns=first_columns, second_columns=second_columns, per=per,
                               average=average, last_column=last_column, last_column_average=last_column_average, columns2=columns2, data2=data2_grouped, third_columns2=third_columns2, computations_columns2=computations_columns2, first_columns2=first_columns2, second_columns2=second_columns2, additional_info=additional_info, ponderate=ponderate)
           
            data3_grouped = None
            if data3 is not None:
                data3_grouped = data3.groupby(list(groups)).get_group(
                    group if len(group) > 1 else group[0])
            rn = row_iteration(sheet,rows_groups, columns, row_number, cnumber, group_data, third_columns, computations_columns, sorting_lists, group=groups, first_columns=first_columns, second_columns=second_columns, per=per,
                               average=average, last_column=last_column, last_column_average=last_column_average, columns3=columns3, data3=data3_grouped, third_columns3=third_columns3, computations_columns3=computations_columns3, first_columns3=first_columns3, second_columns3=second_columns3, 
                               additional_info=additional_info, ponderate=ponderate)
            
            row_number = rn
        # merge last cells
        for i, g in enumerate(group):
            if last_printed[i][0] == g:
                sheet.merge_cells(
                    start_row=last_printed[i][1], start_column=i + 1, end_row=row_number - 2, end_column=i + 1)
                sheet.cell(last_printed[i][1],  i + 1).fill = factors_Fill[i]

def save_workbook(path, workbook, cells_size = NORMAL_WIDTH):
    if "Sheet" in workbook.get_sheet_names():
        std = workbook.get_sheet_by_name('Sheet')
        workbook.remove_sheet(std)
    Adjust_excel_width_height(workbook, cells_size)
    workbook.save(path)
    
def remove_underscore(one_list: List) -> List:
    return [i.replace("_", " ") for i in one_list]

def get_general_cols(rows_groups, general_cols):
    for row in rows_groups:
        if len(rows_groups[row][0]) == 0:
            general_cols.append(row)
        else:
            general_cols += rows_groups[row][0]

def print_basic_sheet(_cfg, rows_groups, name, data, additional_info, groups, workbook, second_column_names, third_columns_names):
    columns = remove_underscore(third_columns_names)
    data = data.round(decimals = 2)
    computations = ["sum"]+ ["mean"]*(len(third_columns_names) - 1)
    Create_excel(workbook.create_sheet(name), rows_groups, third_columns_names, data, columns, computations, _cfg.sorting_lists,
                    second_columns=second_column_names,
                    groups=groups, per = False, average=True, last_column=False, last_column_average=False, additional_info=additional_info)

def Adjust_excel_width_height(workbook: ExcelWriter, multiplier):
        for sheet in workbook.worksheets:
            col_range = sheet[sheet.min_column : sheet.max_column]
            for col in range(1, len(col_range)+1):
                sheet.column_dimensions[get_column_letter(col)].width = WIDTH * multiplier

            row_range = sheet[sheet.min_row : sheet.max_row]
            for row in range(1, len(row_range)+1):
                sheet.row_dimensions[row].height = HEIGHT

def get_groups_add_info(data: DataFrame, row: int, additional_info):
    if type(additional_info) == list:
        additional_info = [ai for ai in additional_info if ai in list(
            data.columns) and ai != row]
        groups = [row] + additional_info
        add_info = len(additional_info)
    else:  # It's a dictionary that indicates in which key it needs to be grouped with
        if row in additional_info:
            groups = [row] + additional_info[row]
            add_info = len(additional_info[row])
        else:
            groups = [row]
            add_info = max(len(additional_info[k]) for k in additional_info)
    return groups, add_info

def columns_alike_our_data(third_columns_names: List[str], second_column_names: List[str], first_column_names: List[str] = None):
    columns = []
    counter_first = 0
    sub_counter_first = 0
    counter_second = 0
    sub_counter_second = 0
    for c in third_columns_names:
        if first_column_names:
            cn = first_column_names[counter_first][0] + \
                second_column_names[counter_second][0] + c
            sub_counter_first += 1
            if sub_counter_first >= first_column_names[counter_first][1]:
                sub_counter_first = 0
                counter_first += 1
        else:
            cn = second_column_names[counter_second][0] + c
        sub_counter_second += 1
        if sub_counter_second >= second_column_names[counter_second][1]:
            sub_counter_second = 0
            counter_second += 1
        columns.append(cn)
    return columns

def print_columns_titles(sheet: ExcelWriter, row: int, column: int, column_names: List[str]):
    for c in column_names:
        sheet.cell(row, column).value = c
        sheet.cell(row, column).font =  FONT_TITLE
        sheet.cell(row, column).fill = titles2Fill
        column += 1

def print_averages_total(sheet: ExcelWriter, row: int, values:List, lable_column: int, values_column: list, average: bool=False, per: bool=False, exception: int=None):
    sheet.cell(row, lable_column).value = "Average" if average else "Total"
    sheet.cell(row, lable_column).font =  FONT_TITLE
    sheet.cell(row, lable_column).fill = orangeFill

    for i, v in enumerate(values):
        if exception and i == exception:  # unicamente ocurre en % Trimmed en Melody_values
            sheet.cell(row, values_column).value = str(round(v * 100, 3)).replace(',','.') + '%'
        else:
            sheet.cell(row, values_column).value = str(v).replace(',','.') if not per else str(v).replace(',','.') + "%"
        values_column += 1

def print_averages_total_column(sheet: ExcelWriter, row: int, column: int, values:List, average: bool=False, per: bool=False):
    sheet.cell(row, column).value = "Average" if average else "Total"
    sheet.cell(row, column).font =  FONT_TITLE
    sheet.cell(row, column).fill = orangeFill
    row += 1
    for v in values:
        sheet.cell(row, column).value = str(v).replace(',','.') if not per else str(v).replace(',','.') + "%"
        row += 1

def write_columns_titles_variable_length(sheet: ExcelWriter, row: int, column: int, column_names: List[tuple], filler = None):
    for i, c in enumerate(column_names):
        if len(column_names) > 2:
            filler=fills_list[i]
        sheet.cell(row, column).value = c[0]
        sheet.cell(row, column).font = FONT_TITLE
        if c[0] != '':
            sheet.cell(row, column).fill = filler
        sheet.cell(row, column).alignment = center
        sheet.merge_cells(start_row=row, start_column=column,
                         end_row=row, end_column=column + c[1] - 1)
        column += c[1]

def split_voices(data: DataFrame):
    # Voices splitting for duetos in Character classification
    data = data.reset_index(drop=True)
    for ind in data.index:
        names = [i for i in data[ROLE].tolist()]
        if '&' in str(names[ind]):
            voices = data[ROLE][ind].split('&')
            pre_data = data.iloc[:ind]
            post_data = data.iloc[ind+1:]
            for i, _ in enumerate(voices):
                line = data.iloc[[ind]]
                line[ROLE] = voices[i]
                line[ARIA_ID] = str(line[ARIA_ID].item()) + \
                    alfa[i] if str(line[ARIA_ID].item()) != '' else ''
                line['Total analysed'] = 1/len(voices)
                line.iloc[:, line.columns.get_loc(
                    "Total analysed")+1:] = line.iloc[:, line.columns.get_loc("Total analysed")+1:].div(len(voices), axis=1)
                pre_data = pre_data.append(line, ignore_index=True)
            data = pre_data.append(post_data).reset_index(drop=True)

    return data

def remove_folder_contents(path: str):
    for filename in os.listdir(path):
        file_path = os.path.join(path, filename)
        if os.path.isfile(file_path):
            os.remove(file_path)
        elif os.path.isdir(file_path):
            remove_folder_contents(file_path)

def print_groups(sheet: ExcelWriter, grouped:DataFrame, row_number: int, column_number: int, columns: list, third_columns: list, computations_columns: list,
                 first_columns: list = None, second_columns: List=None, per: bool=False, average:bool=False, last_column: bool=False,
                 last_column_average: bool=False, additional_info: DataFrame=None, ponderate: bool=False, not_grouped_df:DataFrame=None):
    len_add_info = 0  # Space for the possible column of additional information
    if additional_info:
        len_add_info = additional_info
    if first_columns:
        write_columns_titles_variable_length(
            sheet, row_number, column_number + 1 + len_add_info, first_columns, filler=titles1Fill)
        row_number += 1
    if second_columns:
        write_columns_titles_variable_length(
            sheet, row_number, column_number + 1 + len_add_info, second_columns,filler=titles3Fill)
        row_number += 1
    starting_row = row_number
    print_columns_titles(sheet, row_number, column_number +
                         1 + len_add_info, third_columns)
    row_number += 1
    exception = -1
    total_analysed_column = "Total analysed" in columns
    cnumber = column_number

    columns_values = {c: [] for c in columns}# store each result 

    for s, subgroup in enumerate(grouped):

        cnumber = column_number  
        if type(subgroup[0]) != tuple:  # It has no additional information
            # could be a tuple if we have grouped by more than one element
            sheet.cell(row_number, column_number).value = subgroup[0]
            sheet.cell(row_number, column_number).font =  FONT

        else:
            for g in subgroup[0]:
                sheet.cell(row_number, cnumber).value = g
                sheet.cell(row_number, cnumber).font =  FONT
                cnumber += 1

        subgroup_data = subgroup[1]
        cnumber = column_number + 1 + len_add_info

        total_analysed_row = [1]
        if not_grouped_df is not None:
            if type(subgroup[0]) != tuple:
                cond = not_grouped_df[1][not_grouped_df[0][0]] == subgroup[0]
                not_grouped_information = not_grouped_df[1][cond].drop(
                    not_grouped_df[0] + ['Total analysed'], axis=1)
            else:
                for sb in range(len(subgroup[0])):
                    cond = not_grouped_df[1][not_grouped_df[0]
                                             [sb]] == subgroup[0][sb]
                    not_grouped_information = not_grouped_df[1][cond]
                not_grouped_information = not_grouped_information.drop(
                    not_grouped_df[0] + ['Total analysed'], axis=1)
        else:
            not_grouped_information = None

        for i, c in enumerate(columns):
            column_computation = computations_columns[i]
            if column_computation == 'mean_density':
                value = compute_value(subgroup_data, c, column_computation, total_analysed_row,
                                      not_grouped_information, ponderate)

            elif column_computation == 'mean_texture':
                value = compute_value(subgroup_data, c,  column_computation, total_analysed_row,
                                        not_grouped_information, ponderate)
            else:
                value = compute_value(subgroup_data, c, column_computation, total_analysed_row,
                                      not_grouped_information, ponderate)

            if c == "Total analysed":
                sheet.cell(row_number, cnumber).value = value
                total_analysed_row = subgroup_data[c].to_list()
                sheet.cell(row_number, cnumber).font =  FONT

            if c == interval.ABSOLUTE_INTERVALLIC_TRIM_RATIO:  # EXCEPTION
                sheet.cell(row_number, cnumber).value = str(
                    round(value * 100, 1)) + '%'
                sheet.cell(row_number, cnumber).font =  FONT
                cnumber += 1
                exception = i - 1
            elif not per:
                sheet.cell(row_number, cnumber).value = str(
                    value) + '%' if ponderate and column_computation == "sum" and c != "Total analysed" else str(value).replace(',', '.')
                sheet.cell(row_number, cnumber).font =  FONT
                cnumber += 1
            
            columns_values[c].append(value)
        row_number += 1

    if total_analysed_column:  # We don't need Total analysed from this point
        del columns_values['Total analysed']
        computations_columns = computations_columns[1:]

    last_used_row = row_number
    if per or last_column:  # These are the two conditions in which we need to transpose column_values
        # (change perspective from column to rows)
        
        columns_list = list(columns_values.values())
        keys_columns = list(columns_values.keys())

        rows_values = []
        len_lists = len(columns_list[0])
        for i in range(len_lists):
            rows_values.append(round(sum([lc[i] for x, lc in enumerate(
                columns_list) if "All" not in keys_columns[x]]), 3))

    if per:
        cn = column_number + len_add_info + \
            2 if total_analysed_column else column_number + len_add_info + 1
        for i in range(len(columns_list)):  # Traverse each column's information
            row_number = starting_row
            sum_column = sum(columns_list[i]) if sum(
                columns_list[i]) != 0 else 1
            for j in range(len(columns_list[i])):
                row_number += 1
                # COMPUTE THE HORIZONTAL AVERAGE (average within the present column or row)
                if average:
                    # value = round(
                    #     (columns_list[i][j]/rows_values[j])*100, 3)
                    # values = [i for i in columns_list[i][j] if i! = 0.0]
                    value = round((columns_list[i][j]/rows_values[j])*100, 3)
                else:
                    value = round((columns_list[i][j]/sum_column)*100, 3)
                    
                columns_values[keys_columns[i]][j] = value if str(
                    value) != "nan" else 0  # update the value
                sheet.cell(row_number, cn).value = str(value) + "%" if str(
                    value) != "nan" else 0 # print the value
                sheet.cell(row_number, cn).font =  FONT
            cn += 1

        cnumber = cn

    columns_list = list(columns_values.values()) 

    if per: 
        rows_values = []
        for i in range(len_lists):
            rows_values.append(round(sum([lc[i] for x, lc in enumerate(
                columns_list) if "All" not in keys_columns[x]]), 3))

    if last_column:
        if last_column_average:
            for j, vf in enumerate(rows_values):
                    divisor = (len(columns_list) - len([c for c in keys_columns if "All" in c])- len([col[j] for col in columns_list if col[j] == 0.0]))
                    if divisor != 0:      
                        rows_values[j]=round(vf / divisor, 3)
                    else:
                        rows_values[j]= 0.0
     
        print_averages_total_column(sheet, starting_row, cnumber, rows_values, average=last_column_average,
                                    per=per or ponderate and all(c == "sum" for c in computations_columns))
    
    values=np.asarray(list(columns_values.values()))
    
    for i, c in enumerate(columns_values):
        # In case we have all zeros in a row means that element is not present in the aria so we don't take into account for calculations
        for row in range(0,values.shape[1]):
            if all([str(e)=='0.0' for e in values[:,row]]):
                del columns_values[c][0]

        if average:  
            columns_values[c] = compute_average(columns_values[c], computations_columns[i])
        else:  # total
            columns_values[c] = round(sum(columns_values[c]), 3)


    final_values = list(columns_values.values())
    # Take the last value computed for the last column (average or total)
    if last_column:
        if average:
            final_values.append(
                round(sum(rows_values) / len(rows_values), 3))
        else:
            final_values.append(round(sum(rows_values), 3))
    data_column = column_number + len_add_info + \
        2 if total_analysed_column else column_number + len_add_info + 1
    print_averages_total(sheet, last_used_row, final_values, column_number, data_column, average=average,
                         per=per or ponderate and all(c == "sum" for c in computations_columns), exception=exception)

    return last_used_row + 1, cnumber + 1

def row_iteration(sheet: ExcelWriter, rows_groups: dict, columns: list, row_number: int, column_number: int, data: DataFrame, third_columns: list, computations_columns: List[str], sorting_lists: list, group: list=None, first_columns: list=None, second_columns: list=None, per: bool=False, average: bool=False, last_column: bool=False, last_column_average: bool=False,
                  columns2: list=None, data2: DataFrame=None, third_columns2: list=None, computations_columns2: list=None, first_columns2: list=None, second_columns2: list=None,
                  columns3: list=None, data3: DataFrame=None, third_columns3: list=None, computations_columns3: list=None, first_columns3: list=None, second_columns3: list=None,
                    additional_info: list=[], ponderate: bool =False):
    all_columns = list(data.columns)
    for row in rows_groups:  # Geography, Dramma, Opera, Aria, Label, Composer...    
        if row in all_columns or any(sub in all_columns for sub in rows_groups[row][0]):
            forbiden = [NAME]
            if group != None:
                forbiden = [forbiden_groups[group[i]]
                            for i in range(len(group))]
                forbiden = [item for sublist in forbiden for item in sublist]
            if group == None and row not in forbiden: #it was 'or' and not 'and' before, but change considered neccessary  

                sorting = write_title(sheet, rows_groups, row_number, column_number, row)

                # Write the information depending on the subgroups (ex: Geography -> City, Country)
                if len(rows_groups[row][0]) == 0:  # No subgroups
                    starting_row = row_number
                    data = sort_dataframe(data, row, sorting_lists, sorting)
                    groups_add_info, add_info = get_groups_add_info(
                        data, row, additional_info)
                    row_number, last_column_used = print_groups(sheet, data.groupby(groups_add_info, sort=False), row_number, column_number, columns, third_columns, computations_columns, first_columns, second_columns, per=per,
                                                                average=average, last_column=last_column, last_column_average=last_column_average, additional_info=add_info, ponderate=ponderate, not_grouped_df=(groups_add_info, data[groups_add_info + columns]))
                    if columns2 != None:  # Second subgroup
                        groups_add_info, add_info = get_groups_add_info(
                            data, row, additional_info)
                        if data2 is not None:
                            data2 = sort_dataframe(
                                data2, row, sorting_lists, sorting)
                        row_number, last_column_used = print_groups(sheet, data.groupby(groups_add_info, sort=False) if data2 is None else data2.groupby(groups_add_info, sort=False), starting_row, last_column_used + 1, columns2, third_columns2, computations_columns2, first_columns2,
                                            second_columns2, per=per, average=average, last_column=last_column, last_column_average=last_column_average, additional_info=add_info, ponderate=ponderate, not_grouped_df=(groups_add_info, data[groups_add_info + columns]))
                    
                    if columns3 != None:  # Third subgroup
                        groups_add_info, add_info = get_groups_add_info(
                            data, row, additional_info)    
                        if data3 is not None:
                            data3 = sort_dataframe(
                                data3, row, sorting_lists, sorting)
                        _, _ = print_groups(sheet, data.groupby(groups_add_info, sort=False) if data3 is None else data3.groupby(groups_add_info, sort=False), starting_row, last_column_used + 1, columns3, third_columns3, computations_columns3, first_columns3,
                                            second_columns3, per=per, average=average, last_column=last_column, last_column_average=last_column_average, additional_info=add_info, ponderate=ponderate, not_grouped_df=(groups_add_info, data[groups_add_info + columns]))
                
                else:  # has subgroups, ex: row = Date, subgroups: Year
                    if rows_groups[row][0] == [CHARACTER, ROLE, GENDER]:
                        data_joint = data.copy()
                        data = split_voices(data)
                    for i, subrows in enumerate(rows_groups[row][0]):
                        if (subrows == None or subrows not in forbiden) and subrows in all_columns:
                            if "Tempo" in subrows:
                                data[subrows] = data[subrows].fillna('')
                            starting_row = row_number
                            sort_method = sorting[i]
                            sheet.cell(
                                row_number, column_number).value = subrows
                            sheet.cell(
                                row_number, column_number).fill = greenFill
                            sheet.cell(row_number, column_number).font =  FONT
                            data = sort_dataframe(
                                data, subrows, sorting_lists, sort_method)

                            groups_add_info, add_info = get_groups_add_info(
                                data, subrows, additional_info)
                            row_number, last_column_used = print_groups(sheet, data.groupby(groups_add_info, sort=False), row_number, column_number, columns, third_columns, computations_columns, first_columns, second_columns, per=per,
                                                                        average=average, last_column=last_column, last_column_average=last_column_average, additional_info=add_info, ponderate=ponderate, not_grouped_df=(groups_add_info, data[groups_add_info + columns]))
                            if columns2 != None:  # Second subgroup
                                if "Tempo" in subrows and data2 is not None:
                                    data2[subrows] = data2[subrows].fillna('')
                                if data2 is not None:
                                    data2 = sort_dataframe(
                                        data2, subrows, sorting_lists, sort_method)
                                groups_add_info, add_info = get_groups_add_info(
                                    data, subrows, additional_info)
                                row_number, last_column_used = print_groups(sheet, data.groupby(groups_add_info, sort=False) if data2 is None else data2.groupby(groups_add_info, sort=False), starting_row, last_column_used + 1, columns2, third_columns2, computations_columns2, first_columns2,
                                                    second_columns2, per=per, average=average, last_column=last_column, last_column_average=last_column_average, additional_info=add_info, ponderate=ponderate, not_grouped_df=(groups_add_info, data[groups_add_info + columns]))
                            
                            if columns3 != None:  # Second subgroup
                                if "Tempo" in subrows and data3 is not None:
                                    data3[subrows] = data3[subrows].fillna('')
                                if data3 is not None:
                                    data3 = sort_dataframe(
                                        data3, subrows, sorting_lists, sort_method)
                                groups_add_info, add_info = get_groups_add_info(
                                    data, subrows, additional_info)
                                row_number, last_column_used = print_groups(sheet, data.groupby(groups_add_info, sort=False) if data3 is None else data3.groupby(groups_add_info, sort=False), starting_row, last_column_used + 1, columns3, third_columns3, computations_columns3, first_columns3,
                                                    second_columns3, per=per, average=average, last_column=last_column, last_column_average=last_column_average, additional_info=add_info, ponderate=ponderate, not_grouped_df=(groups_add_info, data[groups_add_info + columns]))

                            row_number += 1
                    if rows_groups[row][0] == [CHARACTER, ROLE, GENDER]:
                        data = copy.deepcopy(data_joint)
                row_number += 1
    return row_number

def write_title(sheet, rows_groups, row_number, column_number, row):
    sheet.cell(row_number, column_number).value = "Per " + row.replace('Aria', '')
    sheet.cell(row_number, column_number).font =  FONT
    sheet.cell(row_number, column_number).fill = YELLOWFILL
    row_number += 1
    sorting = rows_groups[row][1]
    return sorting

def capitalize_instruments(instruments):
    return [instrument[0].upper()+instrument[1:]
                    for instrument in instruments]


def get_excel_name(pre_string, name):
    return pre_string + name + '.xlsx'